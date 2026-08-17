import os
import sys
import tempfile
import unittest
from io import BytesIO
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


class ProductPrdContractTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory(prefix='tmall-products-prd-')
        from app import create_app
        from db import get_db
        self.db = os.path.join(self.tmp.name, 'dashboard.db')
        self.app = create_app({'TESTING': True, 'DATABASE_PATH': self.db})
        self.client = self.app.test_client()
        with get_db(self.db) as c:
            c.executemany('INSERT INTO products (product_id,title,status,tier,style) VALUES (?,?,?,?,?)', [
                ('p-growth', 'Growth', 'active', 'A', 'basic'),
                ('p-decline', 'Decline', 'active', 'B', 'basic'),
            ])
            c.executemany('INSERT INTO monthly_data (product_id,month,payment_amount,refund_amount,ad_spend) VALUES (?,?,?,?,?)', [
                ('p-growth', '2026-04', 200, 10, 20), ('p-decline', '2026-04', 100, 10, 50),
            ])
            c.executemany('INSERT INTO lifecycle_profiles (product_id,recommended_stage,seasonal_attribute) VALUES (?,?,?)', [
                ('p-growth', 'growth', 'spring_summer'), ('p-decline', 'decline', 'stable'),
            ])
            c.execute('INSERT INTO product_actions (id,product_id,purpose_type,purpose_note,action_type,action_detail,target_metric,status,planned_at,observer_window_days) VALUES (?,?,?,?,?,?,?,?,?,?)',
                      ('a1','p-growth','x','x','x','x','payment_amount','pending_execution','2026-04-01',7))
            c.commit()

    def tearDown(self):
        self.tmp.cleanup()

    def test_lifecycle_seasonality_pending_filters_and_derived_fields(self):
        response = self.client.get('/api/products?dim=monthly&period=2026-04&tier=A&lifecycle_stage=growth&seasonality=spring_summer&has_pending_action=true&sort=expense_ratio')
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['availability'], 'available')
        self.assertTrue(payload['requestId'])
        self.assertEqual(payload['evidence_level'], 'full')
        self.assertEqual(payload['evidence'][0]['source'], 'products')
        self.assertIsInstance(payload['freshness'], dict)
        self.assertEqual(payload['data']['total'], 1)
        row = payload['data']['rows'][0]
        self.assertEqual(row['product_id'], 'p-growth')
        self.assertEqual(row['lifecycle_stage'], 'growth')
        self.assertEqual(row['seasonality'], 'spring_summer')
        self.assertTrue(row['has_pending_action'])
        self.assertAlmostEqual(row['expense_ratio'], 0.1)

    def test_products_page_has_explainable_operations_contract(self):
        from pathlib import Path
        page = (Path(ROOT) / 'frontend/ui_demo/pages/products.html').read_text(encoding='utf-8')
        script = (Path(ROOT) / 'frontend/ui_demo/assets/products-live.js').read_text(encoding='utf-8')
        for hook in ('data-products-alert', 'data-products-coverage', 'data-products-issues', 'data-products-health', 'data-products-action'):
            self.assertIn(hook, page)
        self.assertIn('has_data', script)
        self.assertIn('has_pending_action', script)
        self.assertIn('missing facts are not zero', script)

    def test_unknown_sort_is_safe_and_count_uses_filters(self):
        response = self.client.get('/api/products?dim=monthly&period=2026-04&lifecycle_stage=decline&sort=not_a_column')
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['data']['total'], 1)
        self.assertEqual(payload['data']['rows'][0]['product_id'], 'p-decline')

    def test_product_id_sort_does_not_create_ambiguous_sql(self):
        response = self.client.get('/api/products?dim=daily&start=2026-04-01&end=2026-04-30&status=all&sort=product_id&order=asc')
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual([row['product_id'] for row in payload['data']['rows']], ['p-decline', 'p-growth'])

    def test_daily_range_metrics_exclude_previous_period(self):
        from db import get_db
        with get_db(self.db) as c:
            c.executemany(
                'INSERT INTO daily_data (product_id,date,payment_amount,refund_amount,net_sales,ad_spend,ipv,buyers) VALUES (?,?,?,?,?,?,?,?)',
                [
                    ('p-growth', '2026-03-30', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-03-31', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-04-01', 20, 0, 20, 2, 20, 2),
                    ('p-growth', '2026-04-02', 30, 0, 30, 3, 30, 3),
                ],
            )
            c.commit()
        response = self.client.get('/api/products?dim=daily&start=2026-04-01&end=2026-04-02&status=all&product_id=p-growth')
        self.assertEqual(response.status_code, 200)
        row = response.get_json()['data']['rows'][0]
        self.assertEqual(row['payment_amount'], 50)
        self.assertEqual(row['visitors'], 50)
        self.assertEqual(row['ad_spend'], 5)
        self.assertAlmostEqual(row['trend_change'], 1.5)

    def test_net_sales_expense_ratio_and_trend_change_sorts_are_supported(self):
        for sort in ('net_sales', 'expense_ratio', 'trend_change'):
            with self.subTest(sort=sort):
                response = self.client.get(
                    f'/api/products?dim=daily&start=2026-04-01&end=2026-04-30&status=all&sort={sort}&order=desc'
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                response.close()
                self.assertIn('trend_change', payload['data']['rows'][0])

    def test_product_id_filter_is_exact_and_paginates_consistently(self):
        response = self.client.get('/api/products?dim=monthly&period=2026-04&product_id=p-growth')
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['data']['total'], 1)
        self.assertEqual(payload['data']['rows'][0]['product_id'], 'p-growth')

    def test_product_master_rows_without_period_facts_are_partial_evidence(self):
        payload = self.client.get('/api/products?dim=monthly&period=2027-01').get_json()
        self.assertEqual(payload['availability'], 'partial')
        self.assertEqual(payload['evidence_level'], 'partial')
        self.assertIn('product_daily', payload['missing_inputs'])

    def test_export_respects_lifecycle_filter_and_column_order(self):
        response = self.client.post('/api/export', json={
            'type': 'products', 'dim': 'monthly', 'period': '2026-04',
            'lifecycle_stage': 'growth', 'columns': ['title', 'product_id', 'payment_amount'],
        })
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.data)).active
        self.assertEqual([cell.value for cell in next(sheet.iter_rows(max_row=1))], ['商品名称', '商品ID', '销售额'])
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['B2'].value, 'p-growth')

    def test_export_accepts_frontend_metric_column_keys_without_omission(self):
        response = self.client.post('/api/export', json={
            'type': 'products', 'dim': 'monthly', 'period': '2026-04',
            'columns': ['title', 'conversion', 'roi', 'payment_count'],
        })
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.data)).active
        self.assertEqual(sheet.max_column, 4)
        self.assertEqual(sheet.max_row, 3)

    def test_export_includes_lifecycle_pending_and_daily_trend_fields(self):
        from db import get_db
        with get_db(self.db) as c:
            c.executemany(
                'INSERT INTO daily_data (product_id,date,payment_amount,refund_amount,net_sales,ad_spend,ipv,buyers) VALUES (?,?,?,?,?,?,?,?)',
                [
                    ('p-growth', '2026-03-30', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-03-31', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-04-01', 20, 0, 20, 2, 20, 2),
                    ('p-growth', '2026-04-02', 30, 0, 30, 3, 30, 3),
                ],
            )
            c.commit()
        response = self.client.post('/api/export', json={
            'type': 'products', 'dim': 'daily', 'start': '2026-04-01', 'end': '2026-04-02',
            'status': 'all', 'product_id': 'p-growth',
            'columns': ['title', 'lifecycle_stage', 'seasonality', 'has_pending_action', 'trend_change'],
        })
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.data)).active
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ['商品名称', '生命周期阶段', '季节属性', '待办动作', '销售趋势变化'],
        )
        self.assertEqual(sheet[2][1].value, 'growth')
        self.assertEqual(sheet[2][2].value, 'spring_summer')
        self.assertEqual(sheet[2][3].value, 1)
        self.assertAlmostEqual(sheet[2][4].value, 1.5)

    def test_export_uses_daily_range_product_filter_and_sort_order(self):
        from db import get_db
        with get_db(self.db) as c:
            c.executemany(
                'INSERT INTO daily_data (product_id,date,payment_amount,refund_amount,net_sales,ad_spend,ipv,buyers) VALUES (?,?,?,?,?,?,?,?)',
                [
                    ('p-growth', '2026-04-01', 10, 1, 9, 5, 30, 2),
                    ('p-growth', '2026-04-02', 40, 2, 38, 5, 70, 3),
                    ('p-decline', '2026-04-01', 90, 5, 85, 10, 100, 5),
                    ('p-decline', '2026-04-02', 20, 1, 19, 10, 50, 2),
                ],
            )
            c.commit()
        response = self.client.post('/api/export', json={
            'type': 'products', 'dim': 'daily', 'start': '2026-04-01', 'end': '2026-04-02',
            'status': 'all', 'product_id': 'p-growth', 'sort': 'payment_amount', 'order': 'asc',
            'columns': ['product_id', 'payment_amount', 'net_sales', 'visitors'],
        })
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.data)).active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual([cell.value for cell in sheet[2]], ['p-growth', 50, 47, 100])

    def test_export_daily_range_metrics_exclude_previous_period(self):
        from db import get_db
        with get_db(self.db) as c:
            c.executemany(
                'INSERT INTO daily_data (product_id,date,payment_amount,refund_amount,net_sales,ad_spend,ipv,buyers) VALUES (?,?,?,?,?,?,?,?)',
                [
                    ('p-growth', '2026-03-30', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-03-31', 10, 0, 10, 1, 10, 1),
                    ('p-growth', '2026-04-01', 20, 0, 20, 2, 20, 2),
                    ('p-growth', '2026-04-02', 30, 0, 30, 3, 30, 3),
                ],
            )
            c.commit()
        response = self.client.post('/api/export', json={
            'type': 'products', 'dim': 'daily', 'start': '2026-04-01', 'end': '2026-04-02',
            'status': 'all', 'product_id': 'p-growth',
            'sort': 'product_id',
            'columns': ['product_id', 'payment_amount', 'visitors'],
        })
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.data)).active
        self.assertEqual([cell.value for cell in sheet[2]], ['p-growth', 50, 50])


if __name__ == '__main__':
    unittest.main()
