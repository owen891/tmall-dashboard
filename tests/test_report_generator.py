import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.generate_report import collect_metrics, newest_workbook, write_excel


class ReportGeneratorTests(unittest.TestCase):
    def _workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "商品数据"
        ws.append(["统计日期", "商品访客数", "支付买家数", "支付件数", "支付金额", "成功退款金额", "支付新买家数", "支付老买家数", "商品加购人数"])
        ws.append([datetime(2026, 8, 1), 100, 10, 12, 1000, 100, 4, 6, 20])
        ad = wb.create_sheet("推广数据")
        ad.append(["日期", "点击量", "花费", "成交人数", "总成交金额", "总成交笔数", "总购物车数"])
        ad.append([datetime(2026, 8, 1), 50, 100, 8, 500, 9, 11])
        wb.save(path)

    def test_collect_metrics_calculates_requested_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            self._workbook(source)
            dates, matrix, warnings = collect_metrics(source)
            self.assertEqual(dates, ["2026-08-01"])
            self.assertEqual(warnings, [])
            self.assertEqual(matrix[dates[0]]["净销售额"], 900)
            self.assertAlmostEqual(matrix[dates[0]]["商品支付转化率"], 0.1)
            self.assertAlmostEqual(matrix[dates[0]]["投产"], 5)
            self.assertAlmostEqual(matrix[dates[0]]["退款率"], 0.1)
            self.assertAlmostEqual(matrix[dates[0]]["费比"], 100 / 900)

    def test_newest_workbook_ignores_temp_and_reports(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            old = root / "old.xlsx"
            new = root / "new.xlsx"
            old.write_bytes(b"old")
            new.write_bytes(b"new")
            (root / "~$new.xlsx").write_bytes(b"temp")
            (root / "经营数据汇总_20260811.xlsx").write_bytes(b"report")
            self.assertEqual(newest_workbook(root, root), new)

    def test_excel_output_contains_sections_and_dates(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            self._workbook(source)
            dates, matrix, _ = collect_metrics(source)
            output = Path(tmp) / "out.xlsx"
            write_excel(output, dates, matrix, source.name)
            ws = load_workbook(output, data_only=True).active
            labels = [ws.cell(row, 2).value for row in range(3, ws.max_row + 1)]
            self.assertIn("支付金额", labels)
            self.assertEqual(ws.cell(2, 3).value, "2026-08-01")


if __name__ == "__main__":
    unittest.main()
