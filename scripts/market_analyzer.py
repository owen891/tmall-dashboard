# -*- coding: utf-8 -*-
"""
市场数据分析脚本
从3个Excel数据文件读取并分析，输出JSON数据文件
仅包含数据分析逻辑，不包含HTML/CSS/JS生成
"""

import openpyxl
import json
import re
import os
import sys
from collections import defaultdict

# === 常量 ===
NEED_COLORS = {
    u"品类需求": u"#7B3FF2",
    u"适用场景需求": u"#1890ff",
    u"风格需求": u"#fa8c16",
    u"属性需求": u"#52c41a",
    u"其它定制需求": u"#f5222d",
    u"人群需求": u"#eb2f96",
    u"功能属性需求": u"#13c2c2",
    u"品牌需求": u"#2f54eb",
}

DIM_NAMES = [u"品类需求", u"适用场景需求", u"风格需求", u"属性需求", u"其它定制需求", u"人群需求", u"功能属性需求", u"品牌需求"]

# === 全局数据 ===
avg_pop_threshold = 0.0
avg_pop_high = 0.0
avg_pop_low = 0.0

# === Data-driven vocabulary expansion ===
_expanded_patterns = None  # Set by expand_dimension_patterns()


# ============================================================
# 工具函数
# ============================================================

def parse_range(val):
    """处理范围值如 '1000~2500' -> 2500, '5000' -> 5000, '101%' -> 1.01, '-' -> None"""
    if val is None:
        return None
    s = str(val).strip()
    if s == u"-" or s == u"" or s.lower() == u"none":
        return None
    # 处理百分号
    if s.endswith(u"%"):
        try:
            return float(s[:-1].strip().replace(u",", u"")) / 100.0
        except ValueError:
            return None
    if u"~" in s:
        parts = s.split(u"~")
        try:
            return float(parts[-1].strip().replace(u",", u""))
        except (ValueError, IndexError):
            return None
    try:
        return float(s.replace(u",", u""))
    except ValueError:
        return None


def parse_int(val):
    """解析排名值, '-' -> None"""
    if val is None:
        return None
    s = str(val).strip()
    if s == u"-" or s == u"":
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def esc(s):
    """HTML转义"""
    if s is None:
        return u""
    s = str(s)
    s = s.replace(u"&", u"&amp;")
    s = s.replace(u"<", u"&lt;")
    s = s.replace(u">", u"&gt;")
    s = s.replace(u'"', u"&quot;")
    s = s.replace(u"'", u"&#39;")
    return s


def fmt_num(n):
    """格式化数字"""
    if n is None:
        return u"-"
    if n >= 100000000:
        return (u"%.1f" % (n / 100000000.0)) + u"亿"
    if n >= 10000:
        return (u"%.1f" % (n / 10000.0)) + u"万"
    if n >= 1000:
        return u"{:,}".format(int(n))
    return str(int(n))


def fmt_pct(p):
    """格式化百分比"""
    if p is None:
        return u"-"
    return u"%.1f%%" % p


def fmt_range_cn(raw_val):
    """格式化范围值为中文显示"""
    if raw_val is None:
        return u"-"
    s = str(raw_val).strip()
    if s == u"-" or s == u"":
        return u"-"
    if u"~" in s:
        parts = s.split(u"~")
        result = u""
        for i, p in enumerate(parts):
            if i > 0:
                result += u"~"
            pv = parse_range(p)
            if pv is not None:
                result += fmt_num(pv)
            else:
                result += p.strip()
        return result
    pv = parse_range(s)
    if pv is not None:
        return fmt_num(pv)
    return s


def percentile_rank(values, v):
    """Return percentile rank (0~100) of v in values list"""
    if not values:
        return 0.0
    below = sum(1 for x in values if x < v)
    return below / len(values) * 100.0

def percentile_value(values, p):
    """Return the p-th percentile value (p=0~100)"""
    if not values:
        return 0.0
    sorted_data = sorted(values)
    n = len(sorted_data)
    idx = (p / 100.0) * (n - 1)
    lo = int(idx)
    hi = min(lo + 1, n - 1)
    return sorted_data[lo] + (sorted_data[hi] - sorted_data[lo]) * (idx - lo)

def compute_mad(values):
    """Median Absolute Deviation - robust alternative to standard deviation"""
    if not values:
        return 0.0
    sorted_v = sorted(values)
    n = len(sorted_v)
    if n % 2 == 1:
        median = sorted_v[n // 2]
    else:
        median = (sorted_v[n // 2 - 1] + sorted_v[n // 2]) / 2.0
    deviations = [abs(x - median) for x in values]
    sorted_dev = sorted(deviations)
    if n % 2 == 1:
        return sorted_dev[n // 2]
    else:
        return (sorted_dev[n // 2 - 1] + sorted_dev[n // 2]) / 2.0


# ============================================================
# 需求分类
# ============================================================

def expand_dimension_patterns(kw_data_30_local):
    """Data-driven vocabulary expansion: mine high-frequency tokens from actual data"""
    from collections import Counter
    # Get top keywords by search volume
    top_kws = [(kw, pop) for kw, pop, d in kw_data_30_local[:100] if pop and pop > 0]
    if not top_kws:
        return None
    # Simple tokenization: split by non-Chinese chars, filter single chars and stopwords
    stopwords = set(u"的了和与或是在有个一这那用款型版装配件套包本支条适用适合专用通用正品官方旗舰正版价格优惠特价清仓促销礼盒")
    token_freq = Counter()
    for kw, pop in top_kws:
        tokens = re.split(r'[^\u4e00-\u9fffA-Za-z0-9]+', str(kw))
        for t in tokens:
            if len(t) >= 2 and t not in stopwords:
                token_freq[t] += pop
    # Get top 50 high-frequency tokens
    expanded = {}
    for dim_name, dim_color in NEED_COLORS.items():
        # Find tokens that contain any existing seed word
        existing_seeds = set()
        if dim_name == u"品类需求":
            existing_seeds = {u"摆件", u"装饰", u"挂件", u"工艺品", u"花瓶", u"收纳", u"置物架"}
        elif dim_name == u"适用场景需求":
            existing_seeds = {u"客厅", u"卧室", u"桌面", u"玄关", u"办公室", u"书房", u"厨房", u"卫生间"}
        elif dim_name == u"风格需求":
            existing_seeds = {u"中式", u"现代", u"简约", u"复古", u"北欧", u"日式", u"欧式"}
        elif dim_name == u"属性需求":
            existing_seeds = {u"铜", u"水晶", u"陶瓷", u"木质", u"大号", u"小号", u"创意"}
        elif dim_name == u"其它定制需求":
            existing_seeds = {u"定制", u"diy", u"刻字"}
        elif dim_name == u"人群需求":
            existing_seeds = {u"男", u"女", u"儿童", u"老人", u"学生", u"宝宝"}
        elif dim_name == u"功能属性需求":
            existing_seeds = {u"防水", u"保温", u"透气", u"轻便", u"折叠", u"防摔", u"收纳", u"杀菌"}
        elif dim_name == u"品牌需求":
            existing_seeds = set()  # Brand is special, don't auto-expand
        # Find tokens semantically close to existing seeds
        new_tokens = []
        for token, freq in token_freq.most_common(50):
            if token in existing_seeds:
                continue
            for seed in existing_seeds:
                if seed in token or token in seed:
                    new_tokens.append(token)
                    break
        expanded[dim_name] = new_tokens
    return expanded


def classify_need_v2(keyword):
    """8维度需求分类, 返回 ([(dim_name, color), ...], basis_text)"""
    kw = str(keyword).lower()
    dims = []
    basis_parts = []

    # 品类需求
    cat_kw = [u"摆件", u"装饰", u"挂件", u"工艺品", u"花瓶", u"收纳", u"置物架", u"沙漏", u"装饰品", u"小摆件", u"招财猫"]
    cat_matched = []
    for k in cat_kw:
        if k in kw:
            cat_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"品类需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"品类需求"]:
            if ep in kw and ep not in cat_matched:
                cat_matched.append(ep)
    if cat_matched:
        if (u"品类需求", NEED_COLORS[u"品类需求"]) not in dims:
            dims.append((u"品类需求", NEED_COLORS[u"品类需求"]))
            basis_parts.append(u"品类需求（" + u"、" .join(cat_matched) + u"）")

    # 适用场景需求
    scene_kw = [u"客厅", u"卧室", u"桌面", u"玄关", u"办公室", u"书房",
                u"厨房", u"卫生间", u"阳台", u"门", u"窗", u"酒柜",
                u"电视", u"沙发", u"床头", u"茶几", u"电视柜",
                u"家居", u"车载", u"桌上", u"单身公寓", u"电脑桌",
                u"家", u"床头柜", u"写字楼", u"根雕", u"公寓",
                u"开业", u"搬家", u"伴手礼", u"礼物", u"生日礼物",
                u"新房", u"结婚", u"新婚",
                u"礼物", u"旅行", u"会议", u"纪念", u"餐桌",
                u"浴室", u"野餐", u"露营", u"运动", u"健身",
                u"居家", u"出行"]
    scene_matched = []
    for k in scene_kw:
        if k in kw:
            scene_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"适用场景需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"适用场景需求"]:
            if ep in kw and ep not in scene_matched:
                scene_matched.append(ep)
    if scene_matched:
        if (u"适用场景需求", NEED_COLORS[u"适用场景需求"]) not in dims:
            dims.append((u"适用场景需求", NEED_COLORS[u"适用场景需求"]))
            basis_parts.append(u"适用场景需求（" + u"、" .join(scene_matched) + u"）")

    # 风格需求
    style_kw = [u"中式", u"现代", u"简约", u"复古", u"ins", u"北欧",
                u"日式", u"欧式", u"轻奢", u"新中式", u"田园",
                u"古风", u"奶油风", u"侘寂", u"清新", u"微欧",
                u"美式", u"韩式",
                u"可爱", u"文艺", u"浪漫",
                u"彩色", u"纯色", u"格子"]
    style_matched = []
    for k in style_kw:
        if k in kw:
            style_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"风格需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"风格需求"]:
            if ep in kw and ep not in style_matched:
                style_matched.append(ep)
    if style_matched:
        if (u"风格需求", NEED_COLORS[u"风格需求"]) not in dims:
            dims.append((u"风格需求", NEED_COLORS[u"风格需求"]))
            basis_parts.append(u"风格需求（" + u"、" .join(style_matched) + u"）")

    # 属性需求
    attr_kw = [u"铜", u"水晶", u"陶瓷", u"木质", u"大号", u"小号",
               u"创意", u"手工", u"天然", u"玉石", u"树脂",
               u"铁艺", u"玻璃", u"不锈钢", u"石头", u"原石",
               u"琉璃", u"木雕", u"银", u"金", u"玉", u"磨砂",
               u"光滑", u"抗压", u"静音", u"免打孔", u"双面",
               u"电子", u"led", u"夜光", u"小型", u"大型", u"软式",
               u"硬式", u"鸡尾", u"尖锐",
               u"套装", u"加厚", u"迷你", u"单件", u"加大", u"便携",
               u"进口", u"升级", u"整箱", u"独立装",
               u"双层", u"三层", u"加量", u"特大", u"超大", u"超薄",
               u"规格", u"尺寸", u"加长"]
    attr_matched = []
    for k in attr_kw:
        if k in kw:
            attr_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"属性需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"属性需求"]:
            if ep in kw and ep not in attr_matched:
                attr_matched.append(ep)
    if attr_matched:
        if (u"属性需求", NEED_COLORS[u"属性需求"]) not in dims:
            dims.append((u"属性需求", NEED_COLORS[u"属性需求"]))
            basis_parts.append(u"属性需求（" + u"、" .join(attr_matched) + u"）")

    # 其它定制需求
    custom_kw = [u"定制", u"diy", u"刻字", u"personalised", u"来图",
                 u"专属", u"专属定制", u"个性化", u"印刷",
                 u"定做", u"代写", u"开光", u"神像", u"佛",
                 u"财神", u"观音", u"贾宝", u"关公", u"天官赐福",
                 u"神州", u"全民开店", u"十个勤天", u"zarahome"]
    custom_matched = []
    for k in custom_kw:
        if k in kw:
            custom_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"其它定制需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"其它定制需求"]:
            if ep in kw and ep not in custom_matched:
                custom_matched.append(ep)
    if custom_matched:
        if (u"其它定制需求", NEED_COLORS[u"其它定制需求"]) not in dims:
            dims.append((u"其它定制需求", NEED_COLORS[u"其它定制需求"]))
            basis_parts.append(u"其它定制需求（" + u"、" .join(custom_matched) + u"）")

    # 人群需求
    crowd_kw = [u"男生", u"女生", u"儿童", u"老人", u"学生", u"宝宝",
                u"男士", u"女士", u"情侣", u"孕妇", u"婴儿", u"新生儿",
                u"男童", u"女童", u"婴幼儿", u"家用", u"商用",
                u"酒店", u"餐厅", u"办公室", u"母婴", u"老年",
                u"成人", u"青少年"]
    crowd_matched = []
    for k in crowd_kw:
        if k in kw:
            crowd_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"人群需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"人群需求"]:
            if ep in kw and ep not in crowd_matched:
                crowd_matched.append(ep)
    if crowd_matched:
        if (u"人群需求", NEED_COLORS[u"人群需求"]) not in dims:
            dims.append((u"人群需求", NEED_COLORS[u"人群需求"]))
            basis_parts.append(u"人群需求（" + u"、" .join(crowd_matched) + u"）")

    # 功能属性需求
    func_kw = [u"防水", u"保温", u"透气", u"轻便", u"折叠", u"防摔",
               u"净化", u"保鲜", u"加热", u"降噪", u"收纳", u"杀菌",
               u"抗菌", u"吸水", u"速干", u"去污", u"除味", u"护肤",
               u"保湿", u"环保", u"天然", u"有机", u"亲肤",
               u"静音", u"免打孔", u"防滑", u"磁吸", u"充电", u"太阳能",
               u"感应", u"智能", u"遥控"]
    func_matched = []
    for k in func_kw:
        if k in kw:
            func_matched.append(k)
    # Check expanded patterns
    if _expanded_patterns and u"功能属性需求" in _expanded_patterns:
        for ep in _expanded_patterns[u"功能属性需求"]:
            if ep in kw and ep not in func_matched:
                func_matched.append(ep)
    if func_matched:
        if (u"功能属性需求", NEED_COLORS[u"功能属性需求"]) not in dims:
            dims.append((u"功能属性需求", NEED_COLORS[u"功能属性需求"]))
            basis_parts.append(u"功能属性需求（" + u"、" .join(func_matched) + u"）")

    # 品牌需求
    brand_kw = [u"三只松鼠", u"小米", u"华为", u"苹果", u"美的",
                u"飞利浦", u"九牧", u"无印良品", u"德国",
                u"北大伊利", u"泰娜海尔", u"宝宝"]
    brand_matched = []
    for k in brand_kw:
        if k in kw:
            brand_matched.append(k)
    # Brand is special, don't auto-expand
    if brand_matched:
        if (u"品牌需求", NEED_COLORS[u"品牌需求"]) not in dims:
            dims.append((u"品牌需求", NEED_COLORS[u"品牌需求"]))
            basis_parts.append(u"品牌需求（" + u"、" .join(brand_matched) + u"）")

    if not dims:
        basis_text = u"未分类到具体需求维度"
    else:
        basis_text = u"、" .join(basis_parts)

    return dims, basis_text


def classify_keyword_opportunity(kw, pop_30d, ctr_7d, cvr_30d, pop_7d, avg_ctr, avg_cvr):
    """Classify keyword into 5 opportunity categories with scoring.
    Returns (category_name, score, reasoning)
    
    Categories:
    1. 行业必争词 - high search + high CTR
    2. 供给不足蓝海词 - high search + low CTR
    3. 小众高意向蓝海词 - low search + high CVR
    4. 需要关注词 - recent trend anomaly
    5. 常规词 - everything else
    """
    scores = {}
    reasons = []
    
    # Calculate trend ratio (annualized)
    if pop_30d and pop_30d > 0:
        daily_avg_30 = pop_30d / 30.0
        trend_ratio = (pop_7d / 7.0) / daily_avg_30 if daily_avg_30 > 0 else 1.0
    else:
        trend_ratio = 1.0
    
    # 1. 行业必争词
    s = 0
    if pop_30d and pop_30d >= avg_pop_threshold:
        s += 2 if pop_30d >= avg_pop_high else 0
    if avg_ctr and avg_ctr > 0 and ctr_7d is not None:
        if ctr_7d >= avg_ctr * 1.15:
            s += 3
        elif ctr_7d >= avg_ctr:
            s += 2
    scores[u"行业必争词"] = s
    if s >= 3:
        reasons.append(u"搜索量高+点击率高")
    
    # 2. 供给不足蓝海词
    s = 0
    if pop_30d and pop_30d >= avg_pop_low:
        s += 2 if pop_30d >= avg_pop_threshold else 0
    if avg_ctr and avg_ctr > 0 and ctr_7d is not None:
        if ctr_7d < avg_ctr * 0.85:
            s += 3
        elif ctr_7d < avg_ctr * 0.95:
            s += 2
    if avg_cvr and avg_cvr > 0 and cvr_30d is not None:
        if cvr_30d >= avg_cvr:
            s += 1
    scores[u"供给不足蓝海词"] = s
    if s >= 3:
        reasons.append(u"搜索量高+点击率低")
    
    # 3. 小众高意向蓝海词
    s = 0
    if avg_cvr and avg_cvr > 0 and cvr_30d is not None:
        if cvr_30d >= avg_cvr * 1.30:
            s += 3
        elif cvr_30d >= avg_cvr * 1.10:
            s += 2
    if pop_30d and pop_30d < avg_pop_high:
        s += 2 if pop_30d < avg_pop_low else 0
    scores[u"小众高意向蓝海词"] = s
    if s >= 3:
        reasons.append(u"转化率高+搜索量低")
    
    # 4. 需要关注词
    s = 0
    if trend_ratio >= 1.5:
        s += 3
    elif trend_ratio >= 1.15:
        s += 2
    if avg_ctr and avg_ctr > 0 and ctr_7d is not None:
        ctr_dev = ctr_7d - avg_ctr
        if ctr_dev >= 0.10:
            s += 2
        elif ctr_dev >= 0.05:
            s += 1
    scores[u"需要关注词"] = s
    if s >= 2:
        reasons.append(u"近期趋势异动")
    
    # Determine best category
    priority = [u"行业必争词", u"供给不足蓝海词", u"小众高意向蓝海词", u"需要关注词"]
    best_cat = u"常规词"
    best_score = 0
    for cat in priority:
        if scores.get(cat, 0) > best_score:
            best_score = scores[cat]
            best_cat = cat
    
    reasoning = u"；".join(reasons) if reasons else u"常规维护词"
    return best_cat, best_score, reasoning


# ============================================================
# 文件识别
# ============================================================

def _extract_cat_key(fn):
    """从文件名提取品类key用于分组"""
    clean = re.sub(r'^[a-f0-9-]+_', '', fn)
    clean = re.sub(r'\.(xlsx?)$', '', clean, flags=re.IGNORECASE)
    # 搜索排行文件: 提取 "家居饰品-摆件类-装饰摆件" 部分
    m = re.search(r'\u641c\u7d22\u4eba\u6c14-(.+?)(?:-\d{4}|$)', clean)
    if m:
        return m.group(1)
    # 趋势文件: 提取日期范围用于后续交叉匹配
    dates = re.findall(r'(\d{8})', clean)
    if len(dates) >= 2:
        return '__trend__' + dates[0] + '~' + dates[1]
    return clean[:30]


def identify_files(file_paths):
    """自动识别30天/7天/趋势文件，支持多品类文件"""
    # Step 1: 分离搜索文件和趋势文件
    search_files = []
    trend_files = []
    for fp in file_paths:
        fn = os.path.basename(fp)
        if u'趋势分析' in fn:
            trend_files.append(fp)
        elif u'搜索排行' in fn:
            search_files.append(fp)

    # Step 2: 按品类分组搜索文件
    category_groups = {}
    for fp in search_files:
        fn = os.path.basename(fp)
        cat_key = _extract_cat_key(fn)
        if cat_key not in category_groups:
            category_groups[cat_key] = []
        category_groups[cat_key].append(fp)

    # Step 3: 将趋势文件分配到品类组（基于日期重叠）
    for fp in trend_files:
        fn = os.path.basename(fp)
        trend_dates = re.findall(r'(\d{8})', fn)
        if len(trend_dates) < 2:
            continue
        t_start, t_end = int(trend_dates[0]), int(trend_dates[1])
        # 找日期重叠最多的品类组
        best_cat = None
        best_overlap = 0
        for cat_key, group in category_groups.items():
            for gfp in group:
                gfn = os.path.basename(gfp)
                g_dates = re.findall(r'(\d{8})', gfn)
                if len(g_dates) >= 2:
                    g_start, g_end = int(g_dates[0]), int(g_dates[1])
                    overlap = max(0, min(t_end, g_end) - max(t_start, g_start))
                    if overlap > best_overlap:
                        best_overlap = overlap
                        best_cat = cat_key
        if best_cat:
            category_groups[best_cat].append(fp)

    # Step 2: 选择文件数最多的品类组
    best_cat = max(category_groups.keys(), key=lambda k: len(category_groups[k]))
    group = category_groups[best_cat]
    print(u"  检测到品类: " + best_cat + u" (" + str(len(group)) + u" 个文件)")

    # Step 3: 在该组内识别文件类型
    f30 = None
    f7 = None
    ft = None
    for fp in group:
        fn = os.path.basename(fp).lower()
        if u"趋势分析" in fn:
            ft = fp
        elif u"搜索排行" in fn:
            # 支持两种日期格式: YYYYMMDD 和 YYYY-MM-DD
            dates = re.findall(r"(\d{4})-?(\d{2})-?(\d{2})", fn)
            if len(dates) >= 2:
                try:
                    from datetime import datetime
                    d1 = datetime(int(dates[0][0]), int(dates[0][1]), int(dates[0][2]))
                    d2 = datetime(int(dates[1][0]), int(dates[1][1]), int(dates[1][2]))
                    diff = abs((d2 - d1).days)
                    if diff <= 10:
                        f7 = fp
                    else:
                        f30 = fp
                except Exception:
                    f30 = fp
            else:
                f30 = fp
    return f30, f7, ft


def extract_category_and_dates(filename):
    """从文件名提取品类路径和日期范围"""
    dates = re.findall(r"(\d{4})-?(\d{2})-?(\d{2})", filename)
    if dates:
        dates = [d[0] + '-' + d[1] + '-' + d[2] for d in dates]
    # 提取品类: 从文件名中找类似 "家居饰品-摆件类-装饰摆件" 的模式
    # 也可能是文件名开头的部分
    bn = os.path.basename(filename)
    # 去掉日期和常见后缀
    name_part = re.sub(r"\d{4}-?\d{2}-?\d{2}", u"", bn)
    name_part = re.sub(r"\.(xlsx?|csv)$", u"", name_part, flags=re.IGNORECASE)
    # 去掉常见前缀
    name_part = re.sub(r"^[a-f0-9-]+_", u"", name_part)
    name_part = re.sub(r"^\u89c2\u6570[-_]?\u641c\u7d22\u6392\u884c[-_]?\u641c\u7d22\u8bcd[-_]?\u641c\u7d22\u4eba\u6c14[-_]?", u"", name_part)
    name_part = name_part.replace(u"_", u" ").strip()
    # 尝试找品类路径 (A-B-C 模式)
    cat_match = re.search(r"([\u4e00-\u9fff]+(?:-[\u4e00-\u9fff]+)+)", name_part)
    if cat_match:
        cat_path = cat_match.group(1).replace(u"-", u" > ")
        cat_short = cat_path.split(u" > ")[-1] if u" > " in cat_path else cat_path
    else:
        cat_path = name_part.strip()
        cat_short = cat_path
    return cat_path, cat_short, dates


# ============================================================
# 数据加载
# ============================================================

def load_data(f30_path, f7_path, ft_path):
    """加载并解析所有数据，返回 (kw_data_30, kw_data_7, datat, top5_kw, avg_ctr_7d, avg_cvr_30d)"""
    print(u"正在加载30天数据: " + f30_path)
    wb30 = openpyxl.load_workbook(f30_path, read_only=True, data_only=True)
    ws30 = wb30.active
    headers30 = []
    for cell in ws30[1]:
        headers30.append(str(cell.value).strip() if cell.value else u"")
    rows30 = list(ws30.iter_rows(min_row=2, values_only=True))
    wb30.close()

    kw_col = u"搜索词"
    pop_col = u"预估搜索人气"
    pop_raw = u"搜索人气"
    ctr_col = u"点击率"
    cvr_col = u"预估支付转化率"

    # 构建列映射
    col_map_30 = {}
    for i, h in enumerate(headers30):
        col_map_30[h] = i

    kw_data_30 = []
    for row in rows30:
        if row[col_map_30[kw_col]] is None:
            continue
        kw = str(row[col_map_30[kw_col]]).strip()
        if not kw:
            continue
        pop = parse_range(row[col_map_30[pop_col]])
        if pop is None:
            pop = parse_range(row[col_map_30[pop_raw]])
        raw_pop_val = row[col_map_30[pop_raw]] if pop_raw in col_map_30 else None
        ctr = parse_range(row[col_map_30[ctr_col]]) if ctr_col in col_map_30 else None
        cvr = parse_range(row[col_map_30[cvr_col]]) if cvr_col in col_map_30 else None
        d = {
            u"kw": kw,
            u"pop": pop if pop is not None else 0,
            u"pop_raw": raw_pop_val,
            u"ctr": ctr,
            u"cvr": cvr,
        }
        kw_data_30.append((kw, d[u"pop"], d))

    kw_data_30.sort(key=lambda x: x[1], reverse=True)
    print(u"  已加载 %d 个关键词" % len(kw_data_30))

    # 加载7天数据
    print(u"正在加载7天数据: " + f7_path)
    wb7 = openpyxl.load_workbook(f7_path, read_only=True, data_only=True)
    ws7 = wb7.active
    headers7 = []
    for cell in ws7[1]:
        headers7.append(str(cell.value).strip() if cell.value else u"")
    rows7 = list(ws7.iter_rows(min_row=2, values_only=True))
    wb7.close()

    col_map_7 = {}
    for i, h in enumerate(headers7):
        col_map_7[h] = i

    kw_data_7 = {}
    ctr_sum = 0.0
    ctr_count = 0
    for row in rows7:
        if row[col_map_7[kw_col]] is None:
            continue
        kw = str(row[col_map_7[kw_col]]).strip()
        if not kw:
            continue
        pop = parse_range(row[col_map_7[pop_col]])
        if pop is None:
            pop = parse_range(row[col_map_7[pop_raw]])
        kw_data_7[kw] = pop if pop is not None else 0
        ctr_val = parse_range(row[col_map_7[ctr_col]]) if ctr_col in col_map_7 else None
        if ctr_val is not None:
            ctr_sum += ctr_val
            ctr_count += 1

    avg_ctr_7d = ctr_sum / ctr_count if ctr_count > 0 else 0
    print(u"  已加载 %d 个关键词, 平均CTR: %.1f%%" % (len(kw_data_7), avg_ctr_7d))

    # 计算 30天平均CVR
    cvr_sum = 0.0
    cvr_count = 0
    for kw, pop, d in kw_data_30:
        if d[u"cvr"] is not None:
            cvr_sum += d[u"cvr"]
            cvr_count += 1
    avg_cvr_30d = cvr_sum / cvr_count if cvr_count > 0 else 0
    print(u"  30天平均CVR: %.2f%%" % avg_cvr_30d)

    # top5 关键词
    top5_kw = [kw for kw, pop, d in kw_data_30[:5]]

    # 加载趋势数据
    print(u"正在加载趋势数据: " + ft_path)
    wbt = openpyxl.load_workbook(ft_path, read_only=True, data_only=True)
    wst = wbt.active
    headers_t = []
    for cell in wst[1]:
        headers_t.append(str(cell.value).strip() if cell.value else u"")
    rows_t = list(wst.iter_rows(min_row=2, values_only=True))
    wbt.close()

    col_map_t = {}
    for i, h in enumerate(headers_t):
        col_map_t[h] = i

    datat = []
    for row in rows_t:
        rd = {}
        for h, idx in col_map_t.items():
            rd[h] = row[idx] if idx < len(row) else None
        datat.append(rd)

    print(u"  已加载 %d 个商品" % len(datat))

    return kw_data_30, kw_data_7, datat, top5_kw, avg_ctr_7d, avg_cvr_30d


# ============================================================
# 直方图计算
# ============================================================

def compute_histogram(data_list, col_key, bins):
    """计算直方图分布, bins = [(label, lo, hi), ...]
    data_list: [(kw, pop, row_dict), ...] or [row_dict, ...]
    col_key: 'pop' uses item[1], 'ctr'/'cvr' uses item[2][col_key]
    """
    hist = [0] * len(bins)
    for item in data_list:
        if isinstance(item, tuple) and len(item) >= 3:
            if col_key == 'pop':
                val = item[1]
            else:
                val = item[2].get(col_key) if isinstance(item[2], dict) else None
        elif isinstance(item, dict):
            val = item.get(col_key)
        else:
            continue
        if val is None:
            continue
        try:
            val = float(val)
        except (ValueError, TypeError):
            continue
        for i, (label, lo, hi) in enumerate(bins):
            if hi is None:
                if val >= lo:
                    hist[i] += 1
                    break
            else:
                if lo <= val < hi:
                    hist[i] += 1
                    break
    return {u"labels": [b[0] for b in bins], u"hist": hist}


# ============================================================
# 商品数据处理
# ============================================================

def get_all_trend_products(datat, top_n=10):
    """从趋势数据提取商品列表并计算评分"""
    prod_col = u"商品名称"
    prod_id_col = u"商品ID"
    img_col = u"商品图片链接"
    shop_col = u"店铺名称"
    shop_type_col = u"店铺类型"
    rank_cols = [u"排名(第1周)", u"排名(第2周)", u"排名(第3周)", u"排名(第4周)"]
    buyer_cols = [u"支付买家数(第1周)", u"支付买家数(第2周)", u"支付买家数(第3周)", u"支付买家数(第4周)"]
    visitor_cols = [u"访客数(第1周)", u"访客数(第2周)", u"访客数(第3周)", u"访客数(第4周)"]
    kw_tag_col = u"商品关键词"

    products = []
    for row in datat:
        name = str(row.get(prod_col, u"")).strip()
        if not name:
            continue
        pid = str(row.get(prod_id_col, u"")).strip()
        img_url = str(row.get(img_col, u"")).strip()
        shop_name = str(row.get(shop_col, u"")).strip()
        shop_type = str(row.get(shop_type_col, u"")).strip()

        ranks = []
        buyers = []
        visitors = []
        for rc in rank_cols:
            ranks.append(parse_int(row.get(rc)))
        for bc in buyer_cols:
            buyers.append(parse_range(row.get(bc)))
        for vc in visitor_cols:
            visitors.append(parse_range(row.get(vc)))

        # 累计值
        total_buyers = sum([b for b in buyers if b is not None])
        total_visitors = sum([v for v in visitors if v is not None])

        # 平均排名
        valid_ranks = [r for r in ranks if r is not None]
        avg_rank = sum(valid_ranks) / len(valid_ranks) if valid_ranks else 50

        # 关键词标签
        kw_tags_raw = str(row.get(kw_tag_col, u"")).strip()
        kw_tags = []
        if kw_tags_raw and kw_tags_raw != u"None":
            for tag in re.split(r"[\s,;|\u3001]+", kw_tags_raw):
                tag = tag.strip()
                if tag and len(tag) <= 10 and tag not in kw_tags:
                    kw_tags.append(tag)

        products.append({
            u"name": name,
            u"prod_id": pid,
            u"img_url": img_url if img_url and img_url != u"None" else u"",
            u"shop_name": shop_name,
            u"shop_type": shop_type,
            u"ranks": ranks,
            u"buyers": buyers,
            u"visitors": visitors,
            u"total_buyers": total_buyers,
            u"total_visitors": total_visitors,
            u"avg_rank": avg_rank,
            u"kw_tags": kw_tags[:5],
        })

    # === 评分计算 ===
    # 第一遍: 计算原始分
    for p in products:
        valid_ranks = [r for r in p[u"ranks"] if r is not None]
        avg_rank = p[u"avg_rank"] if p[u"avg_rank"] is not None else 100
        # 使用 max(0, 100-avg_rank) 避免高排名商品得负分
        rank_factor = max(0, 100 - avg_rank) * 10
        p[u"raw_pot"] = p[u"total_buyers"] * 0.4 + p[u"total_visitors"] * 0.2 + rank_factor
        p[u"raw_top"] = p[u"total_visitors"] * 0.3 + p[u"total_buyers"] * 0.3 + rank_factor

        # 稳定性评分
        if len(valid_ranks) >= 2:
            p[u"stab_score"] = max(0, 100 - (max(valid_ranks) - min(valid_ranks)) * 0.5)
        else:
            p[u"stab_score"] = 50.0

        # 增长值
        valid_buyers = [b for b in p[u"buyers"] if b is not None]
        if len(valid_buyers) >= 2:
            p[u"growth"] = valid_buyers[-1] - valid_buyers[0]
        elif len(valid_buyers) == 1:
            p[u"growth"] = valid_buyers[0]
        else:
            p[u"growth"] = 0

    # 找到最大值（clamp 到至少 1 避免除零）
    max_pot = max([p[u"raw_pot"] for p in products]) if products else 1
    max_top = max([p[u"raw_top"] for p in products]) if products else 1
    max_pot = max(max_pot, 1)
    max_top = max(max_top, 1)

    # 第二遍: 归一化（clamp 到 0-100）
    for p in products:
        p[u"pot_score"] = max(0, min(100, p[u"raw_pot"] / max_pot * 100))
        p[u"top_score"] = max(0, min(100, p[u"raw_top"] / max_top * 100))
        p[u"overall"] = max(0, min(100, 0.3 * p[u"pot_score"] + 0.3 * p[u"top_score"] + 0.4 * p[u"stab_score"]))

    # Data completeness score
    for p in products:
        check_fields = [u"total_buyers", u"total_visitors", u"avg_rank", u"growth"]
        filled = sum(1 for f in check_fields if p.get(f, 0) != 0 and p.get(f) is not None)
        valid_weeks = sum(1 for r in p[u"ranks"] if r is not None)
        completeness = filled / len(check_fields)
        trend_comp = min(valid_weeks / 4.0, 1.0)
        p[u"completeness"] = round((completeness * 0.6 + trend_comp * 0.4) * 100, 1)

    # 排序
    growth_ranking = sorted(products, key=lambda x: x[u"growth"], reverse=True)[:top_n]
    overall_ranking = sorted(products, key=lambda x: x[u"total_buyers"], reverse=True)[:top_n]
    top_ranking = sorted(products, key=lambda x: x[u"total_visitors"], reverse=True)[:top_n]
    stab_ranking = sorted(products, key=lambda x: x[u"stab_score"], reverse=True)[:top_n]

    return {
        u"growth": growth_ranking,
        u"overall": overall_ranking,
        u"top": top_ranking,
        u"stab": stab_ranking,
    }


# ============================================================
# JSON 数据构建
# ============================================================

def build_analysis_data(f30_path, f7_path, ft_path):
    """主流程：加载数据并构建完整的JSON分析数据"""
    print(u"=" * 60)
    print(u"市场数据分析脚本")
    print(u"=" * 60)

    # 提取品类和日期
    cat_path_30, cat_short_30, dates_30 = extract_category_and_dates(f30_path)
    _, _, dates_7 = extract_category_and_dates(f7_path)
    _, _, dates_t = extract_category_and_dates(ft_path)

    category_path = cat_path_30
    category_short = cat_short_30
    period_30d = dates_30[0] + u" ~ " + dates_30[-1] if len(dates_30) >= 2 else dates_30[0] if dates_30 else u""
    period_7d = dates_7[0] + u" ~ " + dates_7[-1] if len(dates_7) >= 2 else dates_7[0] if dates_7 else u""
    period_trend = dates_t[0] + u" ~ " + dates_t[-1] if len(dates_t) >= 2 else dates_t[0] if dates_t else u""

    print(u"品类: " + category_path)
    print(u"30天期: " + period_30d)
    print(u"7天期: " + period_7d)
    print(u"趋势期: " + period_trend)
    print()

    # 加载数据
    kw_data_30, kw_data_7, datat, top5_kw, avg_ctr_7d, avg_cvr_30d = load_data(f30_path, f7_path, ft_path)
    print()

    # === 设置全局变量 ===
    global avg_pop_threshold, avg_pop_high, avg_pop_low, _expanded_patterns
    all_pops = [pop for kw, pop, d in kw_data_30 if pop and pop > 0]
    if all_pops:
        avg_pop_threshold = percentile_value(all_pops, 50)  # median
        avg_pop_high = percentile_value(all_pops, 75)
        avg_pop_low = percentile_value(all_pops, 30)
    else:
        avg_pop_threshold = avg_pop_high = avg_pop_low = 0
    print(u"  搜索人气百分位: P30=%.0f P50=%.0f P75=%.0f" % (avg_pop_low, avg_pop_threshold, avg_pop_high))

    # Data-driven vocabulary expansion
    _expanded_patterns = expand_dimension_patterns(kw_data_30)
    if _expanded_patterns:
        total_expanded = sum(len(v) for v in _expanded_patterns.values())
        print(u"  词汇扩展: 新增 %d 个模式" % total_expanded)
    print()

    # === 计算 meta ===
    print(u"正在计算元数据...")
    total_keywords = len(kw_data_30)
    top1_kw = kw_data_30[0][0] if kw_data_30 else u""
    top1_pop = kw_data_30[0][1] if kw_data_30 else 0

    meta = {
        u"category_path": category_path,
        u"category_short": category_short,
        u"period_30d": period_30d,
        u"period_7d": period_7d,
        u"period_trend": period_trend,
        u"total_keywords": total_keywords,
        u"avg_ctr_7d": round(avg_ctr_7d, 2),
        u"avg_cvr_30d": round(avg_cvr_30d, 2),
        u"top5_keywords": top5_kw,
    }
    print(u"  元数据计算完成")

    # === 计算 summary ===
    print(u"正在计算摘要统计...")
    up_count = 0
    down_count = 0
    for kw, pop30, d in kw_data_30[:30]:
        pop7 = kw_data_7.get(kw, 0)
        if pop30 > 0 and pop7 > 0:
            pop7_ann = pop7 * 30.0 / 7.0
            change = (pop7_ann - pop30) / pop30 * 100
            if change > 10:
                up_count += 1
            elif change < -10:
                down_count += 1

    # 需求维度统计
    dim_counts = defaultdict(int)
    for kw, pop, d in kw_data_30:
        dims, _ = classify_need_v2(kw)
        for dim_name, _ in dims:
            dim_counts[dim_name] += 1
    top_dim = max(dim_counts.items(), key=lambda x: x[1]) if dim_counts else (u"未知", 0)

    summary = {
        u"growing_count": up_count,
        u"declining_count": down_count,
        u"top_keyword": top1_kw,
        u"top_pop": top1_pop,
        u"main_dimension": top_dim[0],
        u"main_dim_count": top_dim[1],
    }
    print(u"  摘要统计完成")

    # === 计算 keywords ===
    print(u"正在处理关键词数据...")
    keywords = []
    for idx, (kw, pop30, d) in enumerate(kw_data_30[:100]):
        rank = idx + 1
        pop7 = kw_data_7.get(kw, 0)

        # 趋势计算
        change_pct = None
        if pop30 > 0 and pop7 > 0:
            pop7_ann = pop7 * 30.0 / 7.0
            change_pct = round((pop7_ann - pop30) / pop30 * 100, 1)

        # CTR
        ctr_val = d.get(u"ctr")
        ctr_deviation = None
        if ctr_val is not None and avg_ctr_7d > 0:
            ctr_deviation = round(ctr_val - avg_ctr_7d, 1)

        # 需求分类
        dims, basis = classify_need_v2(kw)
        need_tags = [{u"dim": dim_name, u"color": dim_color} for dim_name, dim_color in dims]

        # 机会分类
        opp_cat, opp_score, opp_reason = classify_keyword_opportunity(
            kw, pop30, ctr_val, d.get(u"cvr"), pop7, avg_ctr_7d, avg_cvr_30d)

        keywords.append({
            u"rank": rank,
            u"keyword": kw,
            u"pop_30d": pop30,
            u"pop_7d": pop7,
            u"change_pct": change_pct,
            u"ctr_7d": round(ctr_val, 1) if ctr_val is not None else None,
            u"ctr_deviation": ctr_deviation,
            u"cvr_30d": round(d.get(u"cvr"), 2) if d.get(u"cvr") is not None else None,
            u"need_tags": need_tags,
            u"need_reason": basis,
            u"opportunity_category": opp_cat,
            u"opportunity_score": opp_score,
        })
    print(u"  关键词数据处理完成 (%d 条)" % len(keywords))

    # === 计算 need_stats ===
    print(u"正在计算需求维度统计...")
    dim_stats = defaultdict(lambda: {u"count": 0, u"pop": 0})
    for kw, pop, d in kw_data_30:
        dims, _ = classify_need_v2(kw)
        for dim_name, _ in dims:
            dim_stats[dim_name][u"count"] += 1
            dim_stats[dim_name][u"pop"] += pop

    total_pop_all = sum([v[u"pop"] for v in dim_stats.values()])
    if total_pop_all == 0:
        total_pop_all = 1

    sorted_dims = sorted(dim_stats.items(), key=lambda x: x[1][u"pop"], reverse=True)
    need_stats = {}
    for dn, stats in sorted_dims:
        pct = round(stats[u"pop"] / total_pop_all * 100, 1)
        need_stats[dn] = {
            u"count": stats[u"count"],
            u"total_pop": stats[u"pop"],
            u"pct": pct,
        }
    print(u"  需求维度统计完成")

    # === 计算 dimension_details ===
    print(u"正在计算维度明细...")
    dim_data = defaultdict(list)
    for kw, pop, d in kw_data_30:
        dims, basis = classify_need_v2(kw)
        for dim_name, dim_color in dims:
            dim_data[dim_name].append((kw, pop, basis))

    dimension_details = {}
    for dn in DIM_NAMES:
        items = dim_data.get(dn, [])
        items.sort(key=lambda x: x[1], reverse=True)
        total_pop = sum([p for _, p, _ in items])
        dimension_details[dn] = {
            u"keywords": [{u"keyword": kw, u"pop": pop, u"reason": basis} for kw, pop, basis in items[:30]],
            u"total_pop": total_pop,
            u"count": len(items),
        }
    print(u"  维度明细计算完成")

    # === 计算 histograms ===
    print(u"正在计算直方图数据...")
    pop_bins = [
        (u"600-1200", 600, 1200),
        (u"1200-2500", 1200, 2500),
        (u"2500-5000", 2500, 5000),
        (u"5000-1万", 5000, 10000),
        (u"1-2万", 10000, 20000),
        (u"2-4万", 20000, 40000),
        (u"4-8万", 40000, 80000),
        (u"8-15万", 80000, 150000),
    ]
    ctr_bins = [
        (u"0-5%", 0, 5),
        (u"5-10%", 5, 10),
        (u"10-15%", 10, 15),
        (u"15-20%", 15, 20),
        (u"20-30%", 20, 30),
        (u"30-50%", 30, 50),
        (u"50-100%", 50, 101),
    ]
    cvr_bins = [
        (u"0-0.5%", 0, 0.5),
        (u"0.5-1%", 0.5, 1),
        (u"1-2%", 1, 2),
        (u"2-3%", 2, 3),
        (u"3-5%", 3, 5),
        (u"5-10%", 5, 10),
        (u">10%", 10, None),
    ]

    histograms = {
        u"pop": compute_histogram(kw_data_30, u"pop", pop_bins),
        u"ctr": compute_histogram(kw_data_30, u"ctr", ctr_bins),
        u"cvr": compute_histogram(kw_data_30, u"cvr", cvr_bins),
    }

    # Price distribution (if trend data has price info)
    if datat and len(datat) > 0:
        price_col = None
        for c in [u"价格", u"单价", u"平均价格"]:
            if c in datat[0]:
                price_col = c
                break
        if price_col:
            prices = []
            for row in datat:
                pv = parse_range(row.get(price_col))
                if pv and pv > 0:
                    prices.append(pv)
            if prices:
                min_p, max_p = min(prices), max(prices)
                if min_p < max_p:
                    n_bins = 8
                    step = (max_p - min_p) / n_bins
                    pop_labels_p = []
                    pop_hist_p = []
                    for i in range(n_bins):
                        lo = min_p + i * step
                        hi = lo + step
                        cnt = sum(1 for p in prices if lo <= p < hi)
                        if i == n_bins - 1:
                            cnt += sum(1 for p in prices if p == max_p)
                        pop_labels_p.append(u"\u00a5%.0f~%.0f" % (lo, hi))
                        pop_hist_p.append(cnt)
                    histograms[u"price"] = {u"labels": pop_labels_p, u"hist": pop_hist_p}
    print(u"  直方图数据计算完成")

    # === 计算 rankings ===
    print(u"正在计算商品排名...")
    rankings = get_all_trend_products(datat, top_n=10)

    # 清理排名数据中的内部字段，只保留需要的字段
    clean_rankings = {}
    for rk, prod_list in rankings.items():
        clean_list = []
        for p in prod_list:
            clean_list.append({
                u"name": p[u"name"],
                u"prod_id": p[u"prod_id"],
                u"img_url": p[u"img_url"],
                u"shop_name": p[u"shop_name"],
                u"shop_type": p[u"shop_type"],
                u"ranks": p[u"ranks"],
                u"buyers": p[u"buyers"],
                u"visitors": p[u"visitors"],
                u"total_buyers": p[u"total_buyers"],
                u"total_visitors": p[u"total_visitors"],
                u"growth": p[u"growth"],
                u"pot_score": round(p[u"pot_score"], 1),
                u"top_score": round(p[u"top_score"], 1),
                u"overall": round(p[u"overall"], 1),
                u"stab_score": round(p[u"stab_score"], 1),
                u"completeness": p[u"completeness"],
                u"kw_tags": p[u"kw_tags"],
            })
        clean_rankings[rk] = clean_list
    print(u"  商品排名计算完成")

    # === 组装最终数据 ===
    print(u"正在组装最终数据...")
    data = {
        u"meta": meta,
        u"summary": summary,
        u"keywords": keywords,
        u"need_stats": need_stats,
        u"dimension_details": dimension_details,
        u"histograms": histograms,
        u"rankings": clean_rankings,
    }

    return data


# ============================================================
# 文件查找
# ============================================================

def find_excel_files(directory):
    """在指定目录中查找Excel文件"""
    excel_files = []
    if not os.path.isdir(directory):
        return excel_files
    for fn in os.listdir(directory):
        if fn.lower().endswith((u".xlsx", u".xls")):
            excel_files.append(os.path.join(directory, fn))
    return excel_files


# ============================================================
# 主入口
# ============================================================

def main():
    if len(sys.argv) >= 4:
        # 从命令行参数获取文件路径
        f30_path = sys.argv[1]
        f7_path = sys.argv[2]
        ft_path = sys.argv[3]
    else:
        # 自动检测
        search_dirs = [u"/workspace/.uploads/", u"/workspace/"]
        all_excel = []
        for d in search_dirs:
            all_excel.extend(find_excel_files(d))

        if len(all_excel) < 3:
            print(u"错误: 找不到足够的Excel文件。")
            print(u"请提供3个Excel文件路径作为参数：")
            print(u"  python market_analyzer.py <30天文件> <7天文件> <趋势文件>")
            print(u"或将文件放入 /workspace/.uploads/ 目录")
            sys.exit(1)

        f30, f7, ft = identify_files(all_excel)
        if not f30 or not f7 or not ft:
            print(u"错误: 无法自动识别文件类型。")
            print(u"找到的文件:")
            for f in all_excel:
                print(u"  " + f)
            print(u"请确保文件名包含\u201c搜索排行\u201d和\u201c趋势分析\u201d关键词")
            sys.exit(1)

        f30_path = f30
        f7_path = f7
        ft_path = ft

    # 构建分析数据
    data = build_analysis_data(f30_path, f7_path, ft_path)

    # 确定输出目录和文件名
    output_dir = os.path.dirname(f30_path)
    if not output_dir:
        output_dir = u"/workspace"

    _, cat_short, _ = extract_category_and_dates(f30_path)
    output_filename = cat_short + u"_analysis_data.json"
    output_path = os.path.join(output_dir, output_filename)

    # 写入JSON文件
    print(u"\n正在写入JSON文件: " + output_path)
    with open(output_path, u"w", encoding=u"utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    file_size = os.path.getsize(output_path)
    print(u"\n分析完成!")
    print(u"输出文件: " + output_path)
    print(u"文件大小: " + u"%.1f" % (file_size / 1024.0) + u" KB")
    print(u"=" * 60)

    return output_path


if __name__ == u"__main__":
    main()
