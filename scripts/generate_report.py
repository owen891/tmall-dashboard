#!/usr/bin/env python3
"""Generate the screenshot-style Excel report and a matching HTML dashboard."""
from __future__ import annotations

import argparse
import io
import json
import math
import os
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

SECTIONS = [
    ("经营数据", ["支付金额", "成功退款金额", "退款率", "净销售额", "商品支付转化率", "客单价"]),
    ("流量数据", ["商品访客数", "支付买家数", "支付件数", "支付新买家数", "新客占比", "支付老买家数", "老客占比", "商品加购人数", "加购率"]),
    ("付费数据", ["点击量", "花费", "费比", "成交人数", "总成交金额", "投产", "总成交笔数", "总购物车数"]),
]
PRODUCT_FIELDS = {
    "商品访客数", "支付买家数", "支付件数", "支付新买家数", "支付老买家数", "商品加购人数",
    "支付金额", "成功退款金额",
}
PROMO_FIELDS = {"点击量", "花费", "成交人数", "总成交金额", "总成交笔数", "总购物车数"}
PERCENT_ROWS = {"退款率", "商品支付转化率", "新客占比", "老客占比", "加购率", "费比"}
RATIO_ROWS = {"投产", "客单价"}
PRODUCT_ANALYSIS_FIELDS = [
    "商品访客数", "商品浏览量", "平均停留时长", "商品详情页跳出率", "商品收藏人数",
    "商品加购件数", "商品加购人数", "支付买家数", "支付件数", "支付金额",
    "成功退款金额", "搜索引导支付转化率", "搜索引导支付买家数",
]
PROMOTION_ANALYSIS_FIELDS = [
    "展现量", "点击量", "花费", "点击率", "平均点击花费", "千次展现花费",
    "直接成交金额", "间接成交金额", "总成交金额", "总成交笔数", "点击转化率",
    "投入产出比", "含预售投产比", "总成交成本", "总购物车数", "加购率",
    "收藏宝贝数", "收藏店铺数", "总收藏加购数", "加购成本", "引导访问量",
    "引导访问人数", "引导访问潜客数", "引导访问潜客占比", "成交新客数",
    "成交新客占比", "成交人数",
]


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").strip()
    try:
        number = float(value)
        return number if math.isfinite(number) else 0.0
    except (TypeError, ValueError):
        return 0.0


def _date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip().replace("/", "-")
        for fmt in ("%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def newest_workbook(source_dir: Path, output_dir: Path) -> Path:
    candidates = []
    for path in source_dir.glob("*.xlsx"):
        if path.name.startswith("~$") or path.name.startswith("经营数据汇总_"):
            continue
        if output_dir != source_dir and output_dir in path.parents:
            continue
        candidates.append(path)
    if not candidates:
        raise FileNotFoundError(f"No .xlsx files found in {source_dir}")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _find_header(ws, required: set[str]) -> tuple[int, dict[str, int]] | None:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), 1):
        mapping = {str(v).strip(): i for i, v in enumerate(row) if v not in (None, "")}
        score = len(required.intersection(mapping))
        if score >= max(3, min(len(required), 6)):
            return row_idx, mapping
    return None


def _aggregate(ws, required: set[str], date_name: str, fields: set[str], warnings: list[str]) -> dict[str, dict[str, float]]:
    found = _find_header(ws, required | {date_name})
    if not found:
        warnings.append(f"未找到工作表 {ws.title} 的有效表头")
        return {}
    header_row, mapping = found
    missing = sorted(fields - set(mapping))
    warnings.extend(f"{ws.title} 缺少字段: {name}" for name in missing)
    result: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    date_col = mapping.get(date_name)
    if date_col is None:
        warnings.append(f"{ws.title} 缺少日期字段: {date_name}")
        return {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        day = _date(row[date_col] if date_col < len(row) else None)
        if not day:
            continue
        for field in fields:
            col = mapping.get(field)
            if col is not None and col < len(row):
                result[day][field] += _num(row[col])
    return result


def collect_metrics(source: Path) -> tuple[list[str], dict[str, dict[str, float | None]], list[str]]:
    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        warnings: list[str] = []
        product = {}
        promo = {}
        for ws in wb.worksheets:
            if _find_header(ws, PRODUCT_FIELDS | {"统计日期"}):
                product = _aggregate(ws, PRODUCT_FIELDS, "统计日期", PRODUCT_FIELDS, warnings)
            elif _find_header(ws, PROMO_FIELDS | {"日期"}):
                promo = _aggregate(ws, PROMO_FIELDS, "日期", PROMO_FIELDS, warnings)
    finally:
        wb.close()
    dates = sorted(set(product) | set(promo))
    return _build_matrix(product, promo, warnings)


def _build_matrix(product: dict[str, dict[str, float]], promo: dict[str, dict[str, float]], warnings: list[str]) -> tuple[list[str], dict[str, dict[str, float | None]], list[str]]:
    matrix: dict[str, dict[str, float | None]] = {}
    dates = sorted(set(product) | set(promo))
    for day in dates:
        p = product.get(day, {})
        a = promo.get(day, {})
        visitors = p.get("商品访客数", 0.0)
        buyers = p.get("支付买家数", 0.0)
        payment = p.get("支付金额", 0.0)
        refund = p.get("成功退款金额", 0.0)
        new_buyers = p.get("支付新买家数", 0.0)
        old_buyers = p.get("支付老买家数", 0.0)
        adders = p.get("商品加购人数", 0.0)
        spend = a.get("花费", 0.0)
        gross = a.get("总成交金额", 0.0)
        matrix[day] = {
            "支付金额": payment,
            "成功退款金额": refund,
            "退款率": refund / payment if payment else None,
            "净销售额": payment - refund,
            "商品支付转化率": buyers / visitors if visitors else None,
            "客单价": payment / buyers if buyers else None,
            "商品访客数": visitors,
            "支付买家数": buyers,
            "支付件数": p.get("支付件数", 0.0),
            "支付新买家数": new_buyers,
            "新客占比": new_buyers / buyers if buyers else None,
            "支付老买家数": old_buyers,
            "老客占比": old_buyers / buyers if buyers else None,
            "商品加购人数": adders,
            "加购率": adders / visitors if visitors else None,
            "点击量": a.get("点击量", 0.0),
            "花费": spend,
            "费比": spend / (payment - refund) if payment != refund else None,
            "成交人数": a.get("成交人数", 0.0),
            "总成交金额": gross,
            "投产": gross / spend if spend else None,
            "总成交笔数": a.get("总成交笔数", 0.0),
            "总购物车数": a.get("总购物车数", 0.0),
        }
    return dates, matrix, warnings


def _aggregate_dataframe(df: pd.DataFrame, date_name: str, fields: set[str], warnings: list[str], label: str) -> dict[str, dict[str, float]]:
    columns = {str(c).strip(): c for c in df.columns}
    missing = sorted(fields - set(columns))
    warnings.extend(f"{label} 缺少字段: {name}" for name in missing)
    if date_name not in columns:
        warnings.append(f"{label} 缺少日期字段: {date_name}")
        return {}
    result: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for _, row in df.iterrows():
        day = _date(row.get(columns[date_name]))
        if not day:
            continue
        for field in fields:
            col = columns.get(field)
            if col is not None:
                result[day][field] += _num(row.get(col))
    return result


def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def collect_analysis_directory(source_dir: Path) -> dict[str, list[dict[str, Any]]]:
    """Keep product and promotion grain for the dashboard analysis tabs."""
    catalog: dict[str, dict[str, Any]] = {}
    product_rows: list[dict[str, Any]] = []
    promotion_rows: list[dict[str, Any]] = []

    for path in sorted(source_dir.glob("*.xlsx")):
        try:
            frame = pd.read_excel(path, sheet_name="基准组")
        except Exception:
            continue
        columns = {str(c).strip(): c for c in frame.columns}
        if "宝贝ID" not in columns:
            continue
        for _, row in frame.iterrows():
            product_id = _text(row.get(columns["宝贝ID"]))
            if not product_id:
                continue
            catalog[product_id] = {
                "商品名称": _text(row.get(columns.get("宝贝名称"))),
                "分层": _text(row.get(columns.get("货品成长阶段"))),
                "营销推广IPV": _num(row.get(columns.get("营销推广IPV"))),
                "营销推广消耗": _num(row.get(columns.get("营销推广消耗"))),
                "营销推广ROI": _num(row.get(columns.get("营销推广ROI"))),
                "收加率": _num(row.get(columns.get("收加率"))),
                "复购率": _num(row.get(columns.get("复购率"))),
                "非推广IPV": _num(row.get(columns.get("非推广IPV"))),
                "搜索IPV": _num(row.get(columns.get("搜索IPV"))),
                "推荐IPV": _num(row.get(columns.get("推荐IPV"))),
            }

    for path in sorted(source_dir.glob("*.xls")):
        try:
            frame = pd.read_excel(path, header=4)
        except Exception:
            continue
        columns = {str(c).strip(): c for c in frame.columns}
        if not {"统计日期", "商品ID"}.issubset(columns):
            continue
        for _, row in frame.iterrows():
            day = _date(row.get(columns["统计日期"]))
            product_id = _text(row.get(columns["商品ID"]))
            if not day or not product_id:
                continue
            item: dict[str, Any] = {
                "日期": day,
                "商品ID": product_id,
                "商品名称": _text(row.get(columns.get("商品名称"))) or catalog.get(product_id, {}).get("商品名称", ""),
                "分层": catalog.get(product_id, {}).get("分层", ""),
            }
            for field in PRODUCT_ANALYSIS_FIELDS:
                item[field] = _num(row.get(columns.get(field)))
            item.update({k: v for k, v in catalog.get(product_id, {}).items() if k not in item})
            product_rows.append(item)

    for path in sorted(source_dir.glob("*.zip")):
        try:
            with zipfile.ZipFile(path) as archive:
                for name in archive.namelist():
                    if not name.lower().endswith(".csv"):
                        continue
                    frame = pd.read_csv(io.BytesIO(archive.read(name)), encoding="gb18030")
                    columns = {str(c).strip(): c for c in frame.columns}
                    if "日期" not in columns:
                        continue
                    for _, row in frame.iterrows():
                        day = _date(row.get(columns["日期"]))
                        if not day:
                            continue
                        item = {
                            "日期": day,
                            "场景ID": _text(row.get(columns.get("场景ID"))),
                            "场景名字": _text(row.get(columns.get("场景名字"))),
                            "计划ID": _text(row.get(columns.get("计划ID"))),
                            "计划名字": _text(row.get(columns.get("计划名字"))),
                        }
                        for field in PROMOTION_ANALYSIS_FIELDS:
                            item[field] = _num(row.get(columns.get(field)))
                        promotion_rows.append(item)
        except Exception:
            continue

    return {"products": product_rows, "promotions": promotion_rows}


def collect_metrics_directory(source_dir: Path) -> tuple[list[str], dict[str, dict[str, float | None]], list[str], str]:
    """Merge daily .xls product exports and zipped promotion CSVs from one folder."""
    warnings: list[str] = []
    product: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    promo: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    files_used: list[str] = []
    for path in sorted(source_dir.iterdir()):
        suffix = path.suffix.lower()
        if suffix == ".xls":
            try:
                df = pd.read_excel(path, header=4)
                part = _aggregate_dataframe(df, "统计日期", PRODUCT_FIELDS, warnings, path.name)
                for day, values in part.items():
                    for field, value in values.items():
                        product[day][field] += value
                files_used.append(path.name)
            except Exception as exc:
                warnings.append(f"读取 {path.name} 失败: {exc}")
        elif suffix == ".zip":
            try:
                with zipfile.ZipFile(path) as archive:
                    for name in archive.namelist():
                        if not name.lower().endswith(".csv"):
                            continue
                        raw = archive.read(name)
                        df = pd.read_csv(io.BytesIO(raw), encoding="gb18030")
                        part = _aggregate_dataframe(df, "日期", PROMO_FIELDS, warnings, name)
                        for day, values in part.items():
                            for field, value in values.items():
                                promo[day][field] += value
                        files_used.append(f"{path.name}:{name}")
            except Exception as exc:
                warnings.append(f"读取 {path.name} 失败: {exc}")
        elif suffix == ".xlsx":
            try:
                dates, matrix, local_warnings = collect_metrics(path)
                if dates:
                    warnings.extend(local_warnings)
                    for day in dates:
                        for field in PRODUCT_FIELDS | {"点击量", "花费", "成交人数", "总成交金额", "总成交笔数", "总购物车数"}:
                            if field in matrix[day] and matrix[day][field] is not None:
                                if field in PRODUCT_FIELDS:
                                    product[day][field] += float(matrix[day][field] or 0)
                                else:
                                    promo[day][field] += float(matrix[day][field] or 0)
                    files_used.append(path.name)
            except Exception:
                pass
    if not files_used:
        raise FileNotFoundError(f"No supported source files found in {source_dir}")
    dates, matrix, warnings = _build_matrix(product, promo, warnings)
    return dates, matrix, warnings, ", ".join(files_used)


def write_excel(output: Path, dates: list[str], matrix: dict[str, dict[str, float | None]], source_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "经营数据汇总"
    thin = Side(style="thin", color="222222")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "section": PatternFill("solid", fgColor="BFE7C8"),
        "label": PatternFill("solid", fgColor="C8E8D0"),
        "green": PatternFill("solid", fgColor="C8E8D0"),
        "blue": PatternFill("solid", fgColor="AFC0D3"),
        "cyan": PatternFill("solid", fgColor="C7F0EF"),
        "lime": PatternFill("solid", fgColor="91DF7D"),
        "yellow": PatternFill("solid", fgColor="FFD83D"),
        "red": PatternFill("solid", fgColor="EF4444"),
    }
    ws.cell(1, 1, "数据源")
    ws.cell(1, 2, source_name)
    ws.cell(2, 1, "日期")
    for col, day in enumerate(dates, 3):
        ws.cell(2, col, day)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False
    row = 3
    section_styles = ["green", "cyan", "yellow"]
    for section_idx, (section, metrics) in enumerate(SECTIONS):
        start = row
        for metric in metrics:
            ws.cell(row, 2, metric)
            if metric in {"净销售额", "客单价", "老客占比", "费比"}:
                ws.cell(row, 2).comment = Comment("重点指标", "Codex")
            for col, day in enumerate(dates, 3):
                value = matrix[day].get(metric)
                cell = ws.cell(row, col, value)
                cell.number_format = "0.0%" if metric in PERCENT_ROWS else ("0.0" if metric in RATIO_ROWS else "#,##0.0")
            row += 1
        ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        ws.cell(start, 1, section)
        for rr in range(start, row):
            ws.cell(rr, 1).fill = fills["section"]
            ws.cell(rr, 2).fill = fills[section_styles[section_idx]]
            metric = ws.cell(rr, 2).value
            if metric == "净销售额":
                ws.cell(rr, 2).fill = fills["red"]
                ws.cell(rr, 2).font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
            elif metric == "客单价":
                ws.cell(rr, 2).fill = fills["blue"]
                ws.cell(rr, 2).font = Font(name="Microsoft YaHei", size=12, bold=True)
            elif metric == "老客占比":
                ws.cell(rr, 2).fill = fills["lime"]
                ws.cell(rr, 2).font = Font(name="Microsoft YaHei", size=12, bold=True)
            elif metric == "费比":
                ws.cell(rr, 2).fill = fills["yellow"]
                ws.cell(rr, 2).font = Font(name="Microsoft YaHei", size=12, bold=True)
            for cc in range(1, len(dates) + 3):
                ws.cell(rr, cc).border = border
                ws.cell(rr, cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(rr, cc).font = Font(name="Microsoft YaHei", size=12)
    for cc in range(1, len(dates) + 3):
        ws.cell(2, cc).fill = fills["section"]
        ws.cell(2, cc).border = border
        ws.cell(2, cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, cc).font = Font(name="Microsoft YaHei", size=12, bold=True)
    for cc in range(1, len(dates) + 3):
        ws.cell(1, cc).border = border
    for rr in range(1, row):
        ws.row_dimensions[rr].height = 26
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_dashboard(directory: Path, dates: list[str], matrix: dict[str, dict[str, float | None]], warnings: list[str], source_name: str) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    payload = {"source": source_name, "generated_at": datetime.now().isoformat(timespec="seconds"), "dates": dates, "sections": [{"name": s, "metrics": m} for s, m in SECTIONS], "data": matrix, "warnings": warnings}
    (directory / "data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    data_json = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    html = f'''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>经营数据仪表盘</title><style>
body{{margin:0;background:#eef4ef;color:#17241b;font:14px "Microsoft YaHei",sans-serif}}header{{padding:24px 32px;background:#173b28;color:#fff}}h1{{margin:0 0 6px;font-size:24px}}main{{padding:24px 32px;overflow:auto}}.meta{{color:#cce8d2}}table{{border-collapse:collapse;min-width:1100px;background:#fff;box-shadow:0 4px 18px #173b281c}}th,td{{border:1px solid #abc5b1;padding:10px 12px;text-align:center;white-space:nowrap}}th{{background:#bfe7c8}}td.label{{text-align:left;font-weight:600;background:#c8e8d0}}td.section{{font-weight:700;background:#bfe7c8;writing-mode:vertical-rl;letter-spacing:2px}}.warn{{margin:12px 0;padding:12px;background:#fff4c2;border-left:4px solid #d3a600}}.chart{{margin:18px 0;padding:14px;background:#fff;border:1px solid #abc5b1;min-width:900px}}select{{padding:7px;border-radius:4px;border:1px solid #abc5b1}}
</style></head><body><header><h1>经营数据仪表盘</h1><div class="meta">数据源：{source_name} · 生成时间：{payload['generated_at']}</div></header><main><label>日期筛选 <select id="day"></select></label><div id="warn"></div><div id="chart" class="chart"></div><div id="table"></div></main><script>const P={data_json};const day=document.querySelector('#day');const table=document.querySelector('#table');const fmt=(v,k)=>v==null?'--':(k==='商品支付转化率'||k==='新客占比'||k==='老客占比'||k==='加购率'||k==='费比'?(v*100).toFixed(1)+'%':Number(v).toLocaleString('zh-CN',{{maximumFractionDigits:1}}));function renderChart(days){{const keys=['支付金额','净销售额','总成交金额'];const colors=['#176b3a','#d27b18','#3577b7'];const vals=keys.map(k=>days.map(d=>Number(P.data[d][k]||0)));const max=Math.max(1,...vals.flat());const w=900,h=220,pad=35;let s='<svg viewBox="0 0 '+w+' '+h+'" width="100%" height="220" role="img" aria-label="金额趋势图"><line x1="'+pad+'" y1="'+(h-pad)+'" x2="'+(w-pad)+'" y2="'+(h-pad)+'" stroke="#abc5b1"/>';vals.forEach((arr,i)=>{{const pts=arr.map((v,j)=>((pad+j*(w-2*pad)/Math.max(1,days.length-1))+','+(h-pad-v/max*(h-2*pad))).replace('undefined','0')).join(' ');s+='<polyline fill="none" stroke="'+colors[i]+'" stroke-width="3" points="'+pts+'"/><text x="'+(pad+i*180)+'" y="18" fill="'+colors[i]+'">'+keys[i]+'</text>'}});s+='</svg>';document.querySelector('#chart').innerHTML=s}}function render(){{const d=day.value;let days=d?[d]:P.dates;renderChart(days);let h='<table><tr><th>分类</th><th>指标</th>'+days.map(x=>'<th>'+x+'</th>').join('')+'</tr>';P.sections.forEach(s=>s.metrics.forEach((m,i)=>{{h+='<tr>'+ (i===0?'<td class="section" rowspan="'+s.metrics.length+'">'+s.name+'</td>':'')+'<td class="label">'+m+'</td>'+days.map(x=>'<td>'+fmt(P.data[x][m],m)+'</td>').join('')+'</tr>'}}));table.innerHTML=h+'</table>'}}P.warnings.length&&(document.querySelector('#warn').innerHTML='<div class="warn">字段提示：'+P.warnings.join('；')+'</div>');day.innerHTML='<option value="">全部日期</option>'+P.dates.map(x=>'<option>'+x+'</option>').join('');day.onchange=render;render();</script></body></html>'''
    (directory / "index.html").write_text(html, encoding="utf-8")


def write_dashboard_polished(directory: Path, dates: list[str], matrix: dict[str, dict[str, float | None]], warnings: list[str], source_name: str) -> None:
    """Write the operator-focused visual dashboard; keep the data contract identical to Excel."""
    directory.mkdir(parents=True, exist_ok=True)
    payload = {"source": source_name, "generated_at": datetime.now().isoformat(timespec="seconds"), "dates": dates, "sections": [{"name": s, "metrics": m} for s, m in SECTIONS], "data": matrix, "warnings": warnings}
    (directory / "data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    html = """<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>经营数据看板</title><style>
:root{--ink:#16251b;--muted:#65756a;--paper:#f4f7f2;--line:#b8cdb9;--green:#1e4b31;--mint:#cbe8cf;--mint-2:#e2f1e3;--blue:#b8c8d9;--cyan:#c8efed;--yellow:#ffda4a;--red:#e54b3f;}
*{box-sizing:border-box}body{margin:0;background:var(--paper);color:var(--ink);font:14px/1.45 "Microsoft YaHei",sans-serif}header{background:var(--green);color:#fff;padding:28px clamp(18px,4vw,56px) 22px;border-bottom:5px solid var(--yellow)}.mast{display:flex;justify-content:space-between;gap:20px;align-items:end;max-width:1480px;margin:0 auto}.eyebrow{font-size:11px;letter-spacing:.18em;text-transform:uppercase;color:#b8d8bd}.title{margin:4px 0 0;font-size:30px;letter-spacing:.04em}.meta{color:#cfe6d2;font-size:12px;text-align:right}.shell{max-width:1480px;margin:0 auto;padding:18px clamp(18px,4vw,56px) 40px}.toolbar{display:flex;align-items:center;justify-content:space-between;gap:14px;padding:12px 0 16px;border-bottom:1px solid var(--line)}.toolbar label{color:var(--muted);font-size:12px}.toolbar select{margin-left:8px;padding:8px 30px 8px 10px;border:1px solid var(--line);background:#fff;color:var(--ink);font:inherit}.status{font-size:12px;color:var(--muted)}.kpis{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));margin:18px 0;border:1px solid var(--line);background:#fff}.kpi{padding:16px 18px;border-right:1px solid var(--line)}.kpi:last-child{border-right:0}.kpi:nth-child(1){background:var(--mint)}.kpi:nth-child(2){background:var(--blue)}.kpi:nth-child(3){background:var(--yellow)}.kpi-label{font-size:12px;color:#334d39}.kpi-value{display:block;margin-top:5px;font-size:26px;font-weight:700;letter-spacing:.02em}.viz{display:grid;grid-template-columns:minmax(0,1.6fr) minmax(260px,.6fr);gap:18px;margin-bottom:18px}.panel{background:#fff;border:1px solid var(--line)}.panel-head{display:flex;justify-content:space-between;align-items:center;padding:13px 16px;border-bottom:1px solid var(--line)}.panel-title{font-weight:700}.legend{display:flex;gap:12px;color:var(--muted);font-size:11px}.legend i{display:inline-block;width:9px;height:9px;margin-right:4px}.chart-wrap{padding:12px 16px 8px;min-height:220px}.chart-wrap svg{width:100%;height:210px;display:block}.notes{padding:16px}.notes h2{font-size:14px;margin:0 0 12px}.notes p{margin:9px 0;color:var(--muted);font-size:12px}.warning{padding:10px 12px;background:#fff4c6;border:1px solid #e8cb57;color:#5b4d12;font-size:12px}.report-table{overflow:auto;border:1px solid var(--line);background:#fff}.report-table table{border-collapse:collapse;min-width:1020px;width:100%}.report-table th,.report-table td{border-right:1px solid var(--line);border-bottom:1px solid var(--line);padding:10px 12px;text-align:center;white-space:nowrap}.report-table th{background:var(--green);color:#fff;font-weight:600;position:sticky;top:0;z-index:2}.report-table th:first-child,.report-table th:nth-child(2){background:#163b27}.report-table td.metric{text-align:left;font-weight:600;background:var(--mint-2);position:sticky;left:0;z-index:1}.report-table td.section{font-weight:700;writing-mode:vertical-rl;letter-spacing:.18em;background:var(--mint);position:sticky;left:0;z-index:1}.report-table tr[data-section="付费数据"] td.metric{background:#fff5c9}.report-table tr[data-section="付费数据"] td.section{background:var(--yellow);color:#4b3d00}.report-table tr[data-section="流量数据"] td.metric{background:#e2f7f6}.report-table tr[data-section="流量数据"] td.section{background:var(--cyan)}.report-table td.alert{color:var(--red);font-weight:700}.foot{padding-top:14px;color:var(--muted);font-size:11px}@media(max-width:840px){.mast{display:block}.meta{text-align:left;margin-top:10px}.viz{grid-template-columns:1fr}.kpis{grid-template-columns:1fr}.kpi{border-right:0;border-bottom:1px solid var(--line)}.kpi:last-child{border-bottom:0}.toolbar{align-items:flex-start;flex-direction:column}}
</style></head><body>
<!-- THESIS: A daily operating ledger that makes the requested metric groups scannable in one pass. OWN-WORLD: green paper, hard rules, mint/blue/yellow metric bands. STORY: choose a date, see the financial pulse, then inspect the exact row matrix. FIRST VIEWPORT: masthead, date control, three primary values, trend panel, and the first table rows. FORM: dense operator dashboard, not a card-grid marketing page. FINISH: unreviewed and undocumented is unfinished; this build ends with the finish review, the verdict, and DESIGN.md -->
<header><div class="mast"><div><div class="eyebrow">TMALL / DAILY OPERATIONS</div><h1 class="title">经营数据看板</h1></div><div class="meta">数据源：__SOURCE__<br>生成时间：__GENERATED__</div></div></header>
<main class="shell"><div class="toolbar"><label for="day">查看日期<select id="day" aria-label="选择查看日期"></select></label><div class="status" id="status"></div></div><section class="kpis" id="kpis" aria-label="核心指标"></section><div id="warn"></div><section class="viz"><div class="panel"><div class="panel-head"><span class="panel-title">金额走势</span><span class="legend"><span><i style="background:#1e4b31"></i>支付金额</span><span><i style="background:#d27b18"></i>净销售额</span><span><i style="background:#3577b7"></i>总成交金额</span></span></div><div class="chart-wrap" id="chart"></div></div><aside class="panel notes"><h2>数据说明</h2><p>经营、流量、付费三组指标与 Excel 汇总保持同一口径。</p><p>百分比按日期汇总值计算，客单价与投产按分母不为零时计算。</p><p class="foot">没有明细的日期显示 --，避免用空值制造趋势。</p></aside></section><section class="report-table" id="table"></section><div class="foot">自动生成 · 文件夹最新源表 · 字段提示会显示在页面顶部</div></main>
<script>const P=__DATA__;const day=document.querySelector('#day');const table=document.querySelector('#table');const fmt=(v,k)=>v==null?'--':(['商品支付转化率','新客占比','老客占比','加购率','费比'].includes(k)?(v*100).toFixed(1)+'%':Number(v).toLocaleString('zh-CN',{maximumFractionDigits:1}));function selectedDays(){return day.value?[day.value]:P.dates}function renderKpis(days){const d=days[days.length-1],v=P.data[d]||{};document.querySelector('#kpis').innerHTML=[['净销售额',v['净销售额'],'mint'],['商品访客数',v['商品访客数'],'blue'],['总成交金额',v['总成交金额'],'yellow']].map(x=>'<div class="kpi"><span class="kpi-label">'+x[0]+' · '+d+'</span><strong class="kpi-value">'+fmt(x[1],x[0])+'</strong></div>').join('');document.querySelector('#status').textContent=day.value?'单日视图':'全周期视图 · '+P.dates.length+' 天'}function renderChart(days){const keys=['支付金额','净销售额','总成交金额'],colors=['#1e4b31','#d27b18','#3577b7'],vals=keys.map(k=>days.map(d=>Number(P.data[d][k]||0))),max=Math.max(1,...vals.flat()),w=900,h=210,pad=28;let s='<svg viewBox="0 0 '+w+' '+h+'" role="img" aria-label="金额走势折线图"><line x1="'+pad+'" y1="'+(h-pad)+'" x2="'+(w-pad)+'" y2="'+(h-pad)+'" stroke="#b8cdb9"/>';vals.forEach((arr,i)=>{const pts=arr.map((v,j)=>(pad+j*(w-2*pad)/Math.max(1,days.length-1))+','+(h-pad-v/max*(h-2*pad))).join(' ');s+='<polyline fill="none" stroke="'+colors[i]+'" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="'+pts+'"/>'});s+='</svg>';document.querySelector('#chart').innerHTML=s}function renderTable(days){let h='<table><thead><tr><th>分类</th><th>指标</th>'+days.map(x=>'<th>'+x+'</th>').join('')+'</tr></thead><tbody>';P.sections.forEach(s=>s.metrics.forEach((m,i)=>{h+='<tr data-section="'+s.name+'">'+(i===0?'<td class="section" rowspan="'+s.metrics.length+'">'+s.name+'</td>':'')+'<td class="metric">'+m+'</td>'+days.map(x=>'<td>'+fmt(P.data[x][m],m)+'</td>').join('')+'</tr>'}));table.innerHTML=h+'</tbody></table>'}function render(){const days=selectedDays();renderKpis(days);renderChart(days);renderTable(days)}if(P.warnings.length)document.querySelector('#warn').innerHTML='<div class="warning">字段提示：'+P.warnings.join('；')+'</div>';day.innerHTML='<option value="">全周期</option>'+P.dates.map(x=>'<option>'+x+'</option>').join('');day.onchange=render;render();</script></body></html>"""
    refinement = """<style>
    :root{--paper:#f7f8f5;--line:#d1d9d0;--forest:#244c34;--forest-dark:#183723;--mint:#dcecdf;--mint-soft:#f0f6ef;--blue:#dbe6ef;--cyan:#e0f3f1;--yellow:#f7df78}
    body{background:var(--paper);color:#25332a;font-size:13px}header{padding:22px clamp(18px,4vw,52px) 18px}.title{font-size:26px;letter-spacing:.01em}.eyebrow{font-size:10px}.meta{font-size:11px;color:#d9e6d9}.shell{padding-top:16px}.toolbar{padding-bottom:14px}.range-tabs button{padding:7px 12px;font-size:12px;background:#fbfcfa}.range-tabs button[aria-pressed="true"]{background:var(--forest-dark);border-color:var(--forest-dark)}.status{font-size:11px}.kpis{margin:14px 0;border-color:var(--line);box-shadow:0 1px 0 #fff}.kpi{padding:13px 16px}.kpi:nth-child(1){background:var(--mint)}.kpi:nth-child(2){background:var(--blue)}.kpi:nth-child(3){background:var(--yellow)}.kpi-label{font-size:11px;color:#506157}.kpi-value{font-size:22px}.viz{gap:14px;margin-bottom:14px}.panel{border-color:var(--line)}.panel-head{padding:10px 14px;font-size:13px}.chart{height:190px;padding:10px 14px}.notes{padding:14px}.notes h2{margin-bottom:8px}.notes p{font-size:11px;margin:7px 0}.report-table{border-color:var(--line)}.report-table table{min-width:960px}.report-table th,.report-table td{padding:8px 10px;font-size:12px}.report-table th:first-child{width:58px;min-width:58px}.report-table th:nth-child(2){width:156px;min-width:156px}.report-table td.section{left:0;width:58px;min-width:58px;padding:8px 6px}.report-table td.metric{left:58px;min-width:156px;box-shadow:1px 0 0 var(--line)}.foot{font-size:10px}
    </style>"""
    html = html.replace("</style>", refinement + "</style>", 1)
    html = html.replace("__DATA__", json.dumps(payload, ensure_ascii=False).replace("</", "<\\/"), 1).replace("__SOURCE__", source_name).replace("__GENERATED__", payload["generated_at"])
    (directory / "index.html").write_text(html, encoding="utf-8")


def write_dashboard_with_periods(directory: Path, dates: list[str], matrix: dict[str, dict[str, float | None]], warnings: list[str], source_name: str, analysis: dict[str, list[dict[str, Any]]] | None = None) -> None:
    """Write the final dashboard with accessible day/week/month/range controls."""
    directory.mkdir(parents=True, exist_ok=True)
    payload = {"source": source_name, "generated_at": datetime.now().isoformat(timespec="seconds"), "dates": dates, "sections": [{"name": s, "metrics": m} for s, m in SECTIONS], "data": matrix, "warnings": warnings, "analysis": analysis or {"products": [], "promotions": []}}
    (directory / "data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    html = """<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>经营数据看板</title><style>
:root{--ink:#14251a;--muted:#6a7c70;--paper:#f3f6f0;--line:#b8cbb8;--forest:#214c32;--forest-dark:#153723;--mint:#c9e7cc;--mint-soft:#e8f3e8;--blue:#b9c8d7;--cyan:#c9efed;--yellow:#ffd84c;--amber:#d67819;--sky:#3779b7}*{box-sizing:border-box}body{margin:0;background:var(--paper);color:var(--ink);font:14px/1.45 "Microsoft YaHei",sans-serif}header{background:var(--forest);color:#fff;padding:28px clamp(18px,4vw,56px) 22px}.mast,.shell{max-width:1480px;margin:auto}.mast{display:flex;justify-content:space-between;align-items:end;gap:20px}.eyebrow{font-size:11px;letter-spacing:.16em;color:#bddbbf}.title{margin:3px 0 0;font-size:30px;letter-spacing:.04em}.meta{font-size:12px;text-align:right;color:#d4ead5}.shell{padding:18px clamp(18px,4vw,56px) 40px}.toolbar{display:flex;align-items:center;justify-content:space-between;gap:18px;padding:0 0 18px;border-bottom:1px solid var(--line)}.range-tabs{display:flex;flex-wrap:wrap;gap:0}.range-tabs button{appearance:none;border:1px solid var(--line);border-right:0;background:#fff;color:#506255;font:600 13px "Microsoft YaHei",sans-serif;padding:8px 14px;cursor:pointer}.range-tabs button:first-child{border-radius:4px 0 0 4px}.range-tabs button:last-child{border-right:1px solid var(--line);border-radius:0 4px 4px 0}.range-tabs button:hover{background:var(--mint-soft);color:var(--forest-dark)}.range-tabs button:focus-visible{outline:3px solid #82ac8a;outline-offset:2px}.range-tabs button[aria-pressed="true"]{background:var(--forest);border-color:var(--forest);color:#fff}.status{font-size:12px;color:var(--muted)}.kpis{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));margin:18px 0;border:1px solid var(--line);background:#fff}.kpi{padding:16px 18px;border-right:1px solid var(--line)}.kpi:last-child{border:0}.kpi:nth-child(1){background:var(--mint)}.kpi:nth-child(2){background:var(--blue)}.kpi:nth-child(3){background:var(--yellow)}.kpi-label{display:block;font-size:12px;color:#344b39}.kpi-value{display:block;font-size:26px;margin-top:4px;font-weight:700}.viz{display:grid;grid-template-columns:minmax(0,1.65fr) minmax(250px,.55fr);gap:18px;margin-bottom:18px}.panel{background:#fff;border:1px solid var(--line)}.panel-head{padding:12px 16px;border-bottom:1px solid var(--line);font-weight:700}.chart{height:220px;padding:12px 16px}.chart svg{width:100%;height:100%}.notes{padding:16px}.notes h2{font-size:14px;margin:0 0 12px}.notes p{font-size:12px;color:var(--muted);margin:8px 0}.warning{margin:0 0 18px;padding:10px 12px;border:1px solid #e4c95e;background:#fff4c7;color:#5c4d10;font-size:12px}.report-table{overflow:auto;border:1px solid var(--line);background:#fff}.report-table table{border-collapse:collapse;min-width:1020px;width:100%}.report-table th,.report-table td{padding:10px 12px;text-align:center;white-space:nowrap;border-right:1px solid var(--line);border-bottom:1px solid var(--line)}.report-table th{background:var(--forest-dark);color:#fff;position:sticky;top:0;z-index:3;font-weight:600}.report-table td.metric{text-align:left;font-weight:600;background:var(--mint-soft);position:sticky;left:0;z-index:2}.report-table td.section{font-weight:700;letter-spacing:.16em;writing-mode:vertical-rl;background:var(--mint);position:sticky;left:0;z-index:2}.report-table tr[data-section="流量数据"] td.metric{background:#e4f7f6}.report-table tr[data-section="流量数据"] td.section{background:var(--cyan)}.report-table tr[data-section="付费数据"] td.metric{background:#fff5ca}.report-table tr[data-section="付费数据"] td.section{background:var(--yellow);color:#4f4100}.foot{padding-top:14px;color:var(--muted);font-size:11px}@media(max-width:840px){.mast{display:block}.meta{text-align:left;margin-top:10px}.toolbar{align-items:flex-start;flex-direction:column}.range-tabs button{padding:8px 10px}.kpis,.viz{grid-template-columns:1fr}.kpi{border-right:0;border-bottom:1px solid var(--line)}.kpi:last-child{border-bottom:0}}
</style></head><body><!-- THESIS: A living operations ledger that changes aggregation level without losing the metric hierarchy. OWN-WORLD: green paper, precise borders, and mint/blue/yellow data bands. STORY: select a time lens, scan the business pulse, then inspect the matching matrix. FIRST VIEWPORT: masthead, segmented time controls, financial summary, trend, and grouped table. FORM: dense operator dashboard. FINISH: unreviewed and undocumented is unfinished; this build ends with the finish review, the verdict, and DESIGN.md --><header><div class="mast"><div><div class="eyebrow">TMALL / DAILY OPERATIONS</div><h1 class="title">经营数据看板</h1></div><div class="meta">数据源：__SOURCE__<br>生成时间：__GENERATED__</div></div></header><main class="shell"><div class="toolbar"><div class="range-tabs" role="group" aria-label="时间范围"><button type="button" data-mode="day" aria-pressed="false">按日</button><button type="button" data-mode="week" aria-pressed="false">按周</button><button type="button" data-mode="month" aria-pressed="false">按月</button><button type="button" data-mode="recent7" aria-pressed="true">最近7天</button><button type="button" data-mode="recent30" aria-pressed="false">最近30天</button></div><div class="status" id="status" aria-live="polite"></div></div><section class="kpis" id="kpis" aria-label="核心指标"></section><div id="warn"></div><section class="viz"><div class="panel"><div class="panel-head">金额走势</div><div class="chart" id="chart"></div></div><aside class="panel notes"><h2>数据说明</h2><p>按周、按月会先汇总人数、金额、花费等基础数据，再计算转化率、费比和投产。</p><p>最近 7 天和最近 30 天保留每日明细，方便观察短期趋势。</p><p>没有明细的日期显示 --，不拿空值伪造趋势。</p></aside></section><section class="report-table" id="table"></section><div class="foot">自动生成 · 文件夹最新源表 · 时间口径与 Excel 指标保持一致</div></main><script>const P=__DATA__,buttons=[...document.querySelectorAll('[data-mode]')],raw=['支付金额','成功退款金额','商品访客数','支付买家数','支付件数','支付新买家数','支付老买家数','商品加购人数','点击量','花费','成交人数','总成交金额','总成交笔数','总购物车数'],rates=['商品支付转化率','新客占比','老客占比','加购率','费比'];let mode='recent7';const fmt=(v,k)=>v==null?'--':(rates.includes(k)?(v*100).toFixed(1)+'%':Number(v).toLocaleString('zh-CN',{maximumFractionDigits:1}));function finalise(v){v['净销售额']=v['支付金额']-v['成功退款金额'];v['商品支付转化率']=v['支付买家数']?v['支付买家数']/v['商品访客数']:null;v['客单价']=v['支付买家数']?v['支付金额']/v['支付买家数']:null;v['新客占比']=v['支付买家数']?v['支付新买家数']/v['支付买家数']:null;v['老客占比']=v['支付买家数']?v['支付老买家数']/v['支付买家数']:null;v['加购率']=v['商品访客数']?v['商品加购人数']/v['商品访客数']:null;v['费比']=v['总成交金额']?v['花费']/v['总成交金额']:null;v['投产']=v['花费']?v['总成交金额']/v['花费']:null;return v}function sumGroup(label,days){const v={};raw.forEach(k=>v[k]=days.reduce((n,d)=>n+Number(P.data[d][k]||0),0));return {label,value:finalise(v)}}function weekKey(d){const x=new Date(d+'T00:00:00'),first=new Date(x.getFullYear(),0,1),days=Math.floor((x-first)/86400000);return x.getFullYear()+'年 第'+(Math.ceil((days+first.getDay()+1)/7))+'周'}function series(){if(mode==='day')return P.dates.map(d=>({label:d,value:P.data[d]}));if(mode==='recent7'||mode==='recent30'){const n=mode==='recent7'?7:30;return P.dates.slice(-n).map(d=>({label:d,value:P.data[d]}))}const buckets={};P.dates.forEach(d=>{const k=mode==='week'?weekKey(d):d.slice(0,7);(buckets[k]??=[]).push(d)});return Object.entries(buckets).map(([k,ds])=>sumGroup(k,ds))}function renderKpis(s){const x=s[s.length-1],v=x.value;document.querySelector('#kpis').innerHTML=[['净销售额',v['净销售额']],['商品访客数',v['商品访客数']],['总成交金额',v['总成交金额']]].map(x=>'<div class="kpi"><span class="kpi-label">'+x[0]+' · '+s[s.length-1].label+'</span><strong class="kpi-value">'+fmt(x[1],x[0])+'</strong></div>').join('')}function renderChart(s){const keys=['支付金额','净销售额','总成交金额'],colors=['#214c32','#d67819','#3779b7'],vals=keys.map(k=>s.map(x=>Number(x.value[k]||0))),max=Math.max(1,...vals.flat()),w=900,h=210,p=28;let out='<svg viewBox="0 0 '+w+' '+h+'" role="img" aria-label="金额走势折线图"><line x1="'+p+'" y1="'+(h-p)+'" x2="'+(w-p)+'" y2="'+(h-p)+'" stroke="#b8cbb8"/>';vals.forEach((arr,i)=>{const pts=arr.map((v,j)=>(p+j*(w-2*p)/Math.max(1,s.length-1))+','+(h-p-v/max*(h-2*p))).join(' ');out+='<polyline fill="none" stroke="'+colors[i]+'" stroke-width="3" stroke-linecap="round" points="'+pts+'"/>'});document.querySelector('#chart').innerHTML=out+'</svg>'}function renderTable(s){let h='<table><thead><tr><th>分类</th><th>指标</th>'+s.map(x=>'<th>'+x.label+'</th>').join('')+'</tr></thead><tbody>';P.sections.forEach(g=>g.metrics.forEach((m,i)=>h+='<tr data-section="'+g.name+'">'+(i===0?'<td class="section" rowspan="'+g.metrics.length+'">'+g.name+'</td>':'')+'<td class="metric">'+m+'</td>'+s.map(x=>'<td>'+fmt(x.value[m],m)+'</td>').join('')+'</tr>'));document.querySelector('#table').innerHTML=h+'</tbody></table>'}function render(){const s=series();renderKpis(s);renderChart(s);renderTable(s);document.querySelector('#status').textContent={day:'按日 · '+s.length+' 天',week:'按周 · '+s.length+' 周',month:'按月 · '+s.length+' 月',recent7:'最近 7 天 · 每日明细',recent30:'最近 30 天 · 每日明细'}[mode]}buttons.forEach(b=>b.addEventListener('click',()=>{mode=b.dataset.mode;buttons.forEach(x=>x.setAttribute('aria-pressed',String(x===b)));render()}));if(P.warnings.length)document.querySelector('#warn').innerHTML='<div class="warning">字段提示：'+P.warnings.join('；')+'</div>';render();</script></body></html>"""
    html = html.replace("__DATA__", json.dumps(payload, ensure_ascii=False).replace("</", "<\\/"), 1).replace("__SOURCE__", source_name).replace("__GENERATED__", payload["generated_at"])
    (directory / "index.html").write_text(html, encoding="utf-8")


def refine_dashboard_file(path: Path) -> None:
    """Apply the final spacing/color pass after the generated HTML template is assembled."""
    html = path.read_text(encoding="utf-8")
    html = html.replace(
        '<h1 class="title">经营数据看板</h1></div><div class="meta">',
        '<h1 class="title">经营数据看板</h1></div><nav class="app-nav" aria-label="页面导航"><button class="active" type="button" data-view="overview">总览</button><button type="button" data-view="product">商品分析</button><button type="button" data-view="promotion">推广分析</button></nav><div class="meta">',
        1,
    )
    html = html.replace('<main class="shell">', '<main class="shell" id="overview">', 1)
    html = html.replace('<section class="viz">', '<section class="viz" id="core">', 1)
    analysis_html = '''<section class="analysis-view" id="product-analysis" hidden><div class="analysis-heading"><div><h2>商品分析</h2><p>按商品汇总流量、成交、退款、加购与推广表现</p></div><div class="analysis-heading-tools"><input id="product-search" class="analysis-search" type="search" placeholder="搜索商品名称 / 品牌 / 商品ID" aria-label="搜索商品名称、品牌或商品ID"><span id="product-count"></span></div></div><section class="analysis-kpis" id="product-kpis"></section><section class="analysis-panel analysis-rank-wide"><h3>净销售额排行榜</h3><div class="rank-list" id="product-rank"></div></section><section class="analysis-panel analysis-table analysis-detail-wide"><h3>商品明细 <small>点击表头可排序</small></h3><div id="product-table"></div></section></section><section class="analysis-view" id="promotion-analysis" hidden><div class="analysis-heading"><div><h2>推广分析</h2><p>按计划汇总投放、点击、成交、加购和投产表现</p></div><div class="analysis-heading-tools"><input id="promotion-search" class="analysis-search" type="search" placeholder="搜索计划名称 / 品牌 / 计划ID" aria-label="搜索计划名称、品牌或计划ID"><span id="promotion-count"></span></div></div><section class="analysis-kpis" id="promotion-kpis"></section><section class="analysis-panel analysis-table analysis-table-wide"><h3>计划明细 <small>当前时间范围 · 点击表头可排序</small></h3><div id="promotion-table"></div></section></section>'''
    html = html.replace('<div class="foot">', analysis_html + '<div class="foot">', 1)
    css = """<style>
    :root{--paper:#f7f8f5;--line:#d1d9d0;--forest:#244c34;--forest-dark:#183723;--mint:#dcecdf;--mint-soft:#f0f6ef;--blue:#dbe6ef;--yellow:#f7df78}
    body{background:var(--paper);color:#25332a;font-size:13px}header{padding:22px clamp(18px,4vw,52px) 18px}.title{font-size:26px;letter-spacing:.01em}.eyebrow{font-size:10px}.meta{font-size:11px}.shell{padding-top:16px}.toolbar{padding-bottom:14px}.range-tabs button{padding:7px 12px;font-size:12px;background:#fbfcfa}.range-tabs button[aria-pressed="true"]{background:var(--forest-dark);border-color:var(--forest-dark)}.status{font-size:11px}.kpis{margin:14px 0;border-color:var(--line);box-shadow:0 1px 0 #fff}.kpi{padding:13px 16px}.kpi:nth-child(1){background:var(--mint)}.kpi:nth-child(2){background:var(--blue)}.kpi:nth-child(3){background:var(--yellow)}.kpi-label{font-size:11px;color:#506157}.kpi-value{font-size:22px}.viz{gap:14px;margin-bottom:14px}.panel{border-color:var(--line)}.panel-head{padding:10px 14px;font-size:13px}.chart{height:190px;padding:10px 14px}.notes{padding:14px}.notes h2{margin-bottom:8px}.notes p{font-size:11px;margin:7px 0}.report-table{border-color:var(--line)}.report-table table{min-width:960px}.report-table th,.report-table td{padding:8px 10px;font-size:12px}.report-table th:first-child{width:58px;min-width:58px}.report-table th:nth-child(2){width:156px;min-width:156px}.report-table td.section{left:0;width:58px;min-width:58px;padding:8px 6px}.report-table td.metric{left:58px;min-width:156px;box-shadow:1px 0 0 var(--line)}.foot{font-size:10px}
    </style>"""
    focus_css = """<style>
    .report-table tr.focus-row td.metric{font-weight:800;background:#fff0e8;color:#9e321f}.report-table tr.focus-row td.metric .focus-flag{display:inline-block;margin-right:7px;color:#a33b28;font-size:10px;letter-spacing:.08em}.report-table tr.focus-row td:not(.metric):not(.section){background:#fffaf6}
    </style>"""
    focus_script = r'''<script>
(function(){
  const focus=new Set(['净销售额','客单价','老客占比','费比']);
  const raw=['支付金额','成功退款金额','商品访客数','支付买家数','支付件数','支付新买家数','支付老买家数','商品加购人数','点击量','花费','成交人数','总成交金额','总成交笔数','总购物车数'];
  const percent=new Set(['退款率','商品支付转化率','新客占比','老客占比','加购率','费比']);
  const display=(v,k)=>v==null?'--':(percent.has(k)?(v*100).toFixed(1)+'%':Number(v).toLocaleString('zh-CN',{maximumFractionDigits:1}));
  function finalize(v){v['净销售额']=v['支付金额']-v['成功退款金额'];v['退款率']=v['支付金额']?v['成功退款金额']/v['支付金额']:null;v['商品支付转化率']=v['商品访客数']?v['支付买家数']/v['商品访客数']:null;v['客单价']=v['支付买家数']?v['支付金额']/v['支付买家数']:null;v['新客占比']=v['支付买家数']?v['支付新买家数']/v['支付买家数']:null;v['老客占比']=v['支付买家数']?v['支付老买家数']/v['支付买家数']:null;v['加购率']=v['商品访客数']?v['商品加购人数']/v['商品访客数']:null;v['费比']=v['净销售额']?v['花费']/v['净销售额']:null;v['投产']=v['花费']?v['总成交金额']/v['花费']:null;return v}
  function grouped(){if(mode==='day')return P.dates.map(d=>({label:d,value:finalize({...P.data[d]})}));if(mode==='recent7'||mode==='recent30'){const n=mode==='recent7'?7:30;return P.dates.slice(-n).map(d=>({label:d,value:finalize({...P.data[d]})}))}const buckets={};P.dates.forEach(d=>{const key=mode==='month'?d.slice(0,7):(()=>{const x=new Date(d+'T00:00:00'),first=new Date(x.getFullYear(),0,1),n=Math.floor((x-first)/86400000);return x.getFullYear()+' W'+Math.ceil((n+first.getDay()+1)/7)})();(buckets[key]??=[]).push(d)});return Object.entries(buckets).map(([label,days])=>{const v={};raw.forEach(k=>v[k]=days.reduce((n,d)=>n+Number(P.data[d][k]||0),0));return {label,value:finalize(v)}})}
  function renderFocus(){const series=grouped();const kpi=series[series.length-1].value;document.querySelector('#kpis').innerHTML=[['净销售额',kpi['净销售额']],['商品访客数',kpi['商品访客数']],['总成交金额',kpi['总成交金额']]].map(x=>'<div class="kpi"><span class="kpi-label">'+x[0]+' · '+series[series.length-1].label+'</span><strong class="kpi-value">'+display(x[1],x[0])+'</strong></div>').join('');let table='<table><thead><tr><th>分类</th><th>指标</th>'+series.map(x=>'<th>'+x.label+'</th>').join('')+'</tr></thead><tbody>';P.sections.forEach(group=>group.metrics.forEach((metric,index)=>{const cls=focus.has(metric)?' focus-row':'';const label=focus.has(metric)?'<span class="focus-flag">重点</span>'+metric:metric;table+='<tr class="'+cls.trim()+'" data-section="'+group.name+'">'+(index===0?'<td class="section" rowspan="'+group.metrics.length+'">'+group.name+'</td>':'')+'<td class="metric">'+label+'</td>'+series.map(x=>'<td>'+display(x.value[metric],metric)+'</td>').join('')+'</tr>'}));document.querySelector('#table').innerHTML=table+'</tbody></table>';document.querySelector('#status').textContent={day:'按日 · '+series.length+' 天',week:'按周 · '+series.length+' 周',month:'按月 · '+series.length+' 月',recent7:'最近 7 天 · 每日明细',recent30:'最近 30 天 · 每日明细'}[mode]}
  render=renderFocus;renderFocus();
})();
</script>'''
    core_css = """<style>
    .kpis{grid-template-columns:repeat(4,minmax(0,1fr))}.kpi.core-kpi{background:#fff;border-right:1px solid var(--line);border-top:3px solid #b7d9bc}.kpi.core-kpi:nth-child(2){border-top-color:#aebfd2}.kpi.core-kpi:nth-child(3){border-top-color:#91c567}.kpi.core-kpi:nth-child(4){border-top-color:#e0b631}.core-chart{height:330px;padding:12px 14px}.core-chart svg{width:100%;height:100%;display:block}.core-chart .axis{fill:#77877d;font-size:11px}.core-chart .series-name{fill:#32483a;font-size:12px;font-weight:700}.core-chart .last-value{font-size:11px;font-weight:700}@media(max-width:840px){.kpis{grid-template-columns:repeat(2,minmax(0,1fr))}.kpi.core-kpi:nth-child(2){border-right:0}.core-chart{height:360px}}
    </style>"""
    time_css = """<style>
    .toolbar{position:relative;display:block;padding:0 0 15px}.time-line{display:flex;align-items:center;gap:10px;color:#557064;font-size:12px;margin-bottom:9px}.time-line strong{font-weight:500;color:#587a9b}.time-picker{display:flex;align-items:center;justify-content:space-between;gap:10px}.time-group{display:flex;align-items:center;border:1px solid #d5ded6;background:#f6f8f5;border-radius:10px;padding:3px}.time-group button{border:0;background:transparent;color:#324a3a;font:600 12px "Microsoft YaHei",sans-serif;padding:7px 11px;cursor:pointer;border-radius:7px}.time-group button:hover{background:#e8f2e8}.time-group button[aria-pressed="true"]{background:#fff;color:#8a2ed1;box-shadow:0 1px 4px #1a39231c}.time-divider{width:1px;height:22px;background:#d5ded6;margin:0 2px}.time-custom{position:relative}.calendar-trigger{display:flex;align-items:center;gap:5px}.calendar-trigger .calendar-icon{font-size:15px;line-height:1}.calendar-popover{position:absolute;right:0;top:calc(100% + 8px);z-index:20;width:620px;background:#fff;border:1px solid #d8e1d9;border-radius:12px;box-shadow:0 16px 35px #18372324;padding:12px;display:none}.calendar-popover.open{display:block}.calendar-head{display:grid;grid-template-columns:30px 1fr 30px;align-items:center;text-align:center;margin-bottom:6px}.calendar-head strong{font-size:18px;color:#18231d}.calendar-nav{border:0;background:transparent;color:#60776b;font-size:22px;cursor:pointer}.calendar-panes{display:grid;grid-template-columns:1fr 1fr;gap:18px}.calendar-pane{min-width:0}.calendar-pane h3{margin:6px 0 8px;text-align:center;font-size:14px}.calendar-week,.calendar-grid{display:grid;grid-template-columns:repeat(7,1fr);text-align:center}.calendar-week span{color:#89988d;font-size:11px;padding:4px 0}.calendar-grid button{border:0;background:transparent;color:#33483a;height:30px;border-radius:7px;cursor:pointer}.calendar-grid button:hover{background:#f0e6ff;color:#8a2ed1}.calendar-grid button.muted{color:#b4bdb5}.calendar-grid button.selected{background:#8b32d5;color:#fff}.calendar-grid button.in-range{background:#f1e6fb;color:#7132a9;border-radius:0}.calendar-footer{display:flex;justify-content:space-between;align-items:center;border-top:1px solid #e1e7e1;margin-top:10px;padding-top:10px;color:#60776b;font-size:12px}.calendar-footer button{border:1px solid #cbd8cc;background:#fff;color:#35523f;border-radius:6px;padding:6px 12px;cursor:pointer}@media(max-width:840px){.time-picker{display:block}.time-group{width:100%;justify-content:space-between}.time-group button{padding:7px 8px}.calendar-popover{position:fixed;left:12px;right:12px;top:140px;width:auto}.calendar-panes{gap:8px}.calendar-pane:nth-child(2){display:none}}
    </style>"""
    time_script = r'''<script>
(function(){
  const toolbar=document.querySelector('.toolbar'), old=toolbar.querySelector('.range-tabs');
  const modes=[['yesterday','昨天'],['7','7天'],['30','30天'],['60','60天'],['custom','自定义']];
  let range='7',granularity='day',anchor=new Date(P.dates[P.dates.length-1]+'T00:00:00'),customStart=P.dates[0],customEnd=P.dates[P.dates.length-1],calendarOpen=false,selectingStart=true;
  toolbar.innerHTML='<div class="time-line"><span>统计时间</span><strong id="time-label"></strong></div><div class="time-picker"><div class="time-group" id="range-group" role="group" aria-label="时间范围">'+modes.map(x=>'<button type="button" data-range="'+x[0]+'" aria-pressed="'+(x[0]==='7')+'">'+x[1]+'</button>').join('')+'<span class="time-divider"></span><button type="button" data-gran="day" aria-pressed="true">日</button><button type="button" data-gran="week" aria-pressed="false">周</button><button type="button" data-gran="month" aria-pressed="false">月</button><span class="time-custom"><button type="button" class="calendar-trigger" data-range="custom" aria-expanded="false"><span class="calendar-icon" aria-hidden="true">▦</span>自定义</button><div class="calendar-popover" id="calendar"></div></span></div><div class="status" id="status" aria-live="polite"></div></div>';
  function datesForRange(){if(range==='yesterday')return P.dates.slice(-1);if(range==='7')return P.dates.slice(-7);if(range==='30')return P.dates.slice(-30);if(range==='60')return P.dates.slice(-60);return P.dates.filter(d=>d>=customStart&&d<=customEnd)}
  function keyWeek(d){const x=new Date(d+'T00:00:00'),first=new Date(x.getFullYear(),0,1),n=Math.floor((x-first)/86400000);return x.getFullYear()+' W'+Math.ceil((n+first.getDay()+1)/7)}
  function seriesForRange(){const days=datesForRange();if(granularity==='day')return days.map(d=>({label:d,value:finalize({...P.data[d]})}));const buckets={};days.forEach(d=>{const k=granularity==='month'?d.slice(0,7):keyWeek(d);(buckets[k]??=[]).push(d)});return Object.entries(buckets).map(([label,items])=>{const v={};raw.forEach(k=>v[k]=items.reduce((n,d)=>n+Number(P.data[d][k]||0),0));return {label,value:finalize(v)}})}
  function calendarMonth(year,month){const first=new Date(year,month,1),start=new Date(year,month,1-first.getDay()),cells='日一二三四五六'.split('').map(x=>'<span>'+x+'</span>').join('');let days='';for(let i=0;i<42;i++){const d=new Date(start);d.setDate(start.getDate()+i);const iso=d.toISOString().slice(0,10),muted=d.getMonth()!==month,selected=iso===customStart||iso===customEnd,inRange=iso>customStart&&iso<customEnd;days+='<button type="button" class="'+(muted?'muted ':'')+(selected?'selected ':'')+(inRange?'in-range':'')+'" data-date="'+iso+'">'+d.getDate()+'</button>'}return '<div class="calendar-pane"><h3>'+year+'年 '+(month+1)+'月</h3><div class="calendar-week">'+cells+'</div><div class="calendar-grid">'+days+'</div></div>'}
  function renderCalendar(){const y=anchor.getFullYear(),m=anchor.getMonth();document.querySelector('#calendar').innerHTML='<div class="calendar-head"><button type="button" class="calendar-nav" data-cal="prev" aria-label="上一个月">‹</button><strong>选择日期范围</strong><button type="button" class="calendar-nav" data-cal="next" aria-label="下一个月">›</button></div><div class="calendar-panes">'+calendarMonth(y,m)+calendarMonth(m===11?y+1:y,m===11?0:m+1)+'</div><div class="calendar-footer"><span>'+customStart+' ~ '+customEnd+'</span><button type="button" data-cal="done">完成</button></div>';document.querySelectorAll('[data-date]').forEach(b=>b.addEventListener('click',()=>{const d=b.dataset.date;if(selectingStart){customStart=d;customEnd=d;selectingStart=false}else{if(d<customStart){customEnd=customStart;customStart=d}else{customEnd=d}selectingStart=true}renderCalendar();renderView()}));document.querySelectorAll('[data-cal]').forEach(b=>b.addEventListener('click',()=>{if(b.dataset.cal==='prev')anchor.setMonth(anchor.getMonth()-1);if(b.dataset.cal==='next')anchor.setMonth(anchor.getMonth()+1);if(b.dataset.cal==='done'){calendarOpen=false;document.querySelector('#calendar').classList.remove('open')}renderCalendar()}))}
  function renderView(){const s=seriesForRange();document.querySelector('#time-label').textContent=(datesForRange()[0]||'--')+' ~ '+(datesForRange().slice(-1)[0]||'--');document.querySelector('#kpis').innerHTML=[['净销售额',s[s.length-1]?.value['净销售额']],['商品访客数',s[s.length-1]?.value['商品访客数']],['总成交金额',s[s.length-1]?.value['总成交金额']]].map(x=>'<div class="kpi"><span class="kpi-label">'+x[0]+' · '+(s[s.length-1]?.label||'--')+'</span><strong class="kpi-value">'+display(x[1],x[0])+'</strong></div>').join('');let table='<table><thead><tr><th>分类</th><th>指标</th>'+s.map(x=>'<th>'+x.label+'</th>').join('')+'</tr></thead><tbody>';P.sections.forEach(g=>g.metrics.forEach((metric,index)=>{const cls=focus.has(metric)?' focus-row':'';const label=focus.has(metric)?'<span class="focus-flag">重点</span>'+metric:metric;table+='<tr class="'+cls.trim()+'" data-section="'+g.name+'">'+(index===0?'<td class="section" rowspan="'+g.metrics.length+'">'+g.name+'</td>':'')+'<td class="metric">'+label+'</td>'+s.map(x=>'<td>'+display(x.value[metric],metric)+'</td>').join('')+'</tr>'}));document.querySelector('#table').innerHTML=table+'</tbody></table>';document.querySelector('#status').textContent=(granularity==='day'?'按日':granularity==='week'?'按周':'按月')+' · '+s.length+' 个时间段';renderChart(s)}
  document.querySelectorAll('[data-range]').forEach(b=>b.addEventListener('click',()=>{range=b.dataset.range;if(range==='custom'){calendarOpen=!calendarOpen;document.querySelector('#calendar').classList.toggle('open',calendarOpen);document.querySelector('.calendar-trigger').setAttribute('aria-expanded',String(calendarOpen));renderCalendar()}else{document.querySelector('#calendar').classList.remove('open');renderView()}document.querySelectorAll('[data-range]').forEach(x=>x.setAttribute('aria-pressed',String(x===b)))}));document.querySelectorAll('[data-gran]').forEach(b=>b.addEventListener('click',()=>{granularity=b.dataset.gran;document.querySelectorAll('[data-gran]').forEach(x=>x.setAttribute('aria-pressed',String(x===b)));renderView()}));renderCalendar();renderView();
})();
</script>'''
    core_logic = r'''function renderCoreChart(s){const keys=['净销售额','客单价','老客占比','费比'],colors=['#2e6b3f','#536f8d','#6eaa45','#bd8b20'],w=960,rowH=72,left=132,right=74,top=18,bottom=38,h=top+rowH*keys.length+bottom;let svg='<svg viewBox="0 0 '+w+' '+h+'" role="img" aria-label="核心指标日期趋势图">';keys.forEach((key,i)=>{const values=s.map(x=>Number(x.value[key]??0)),min=Math.min(...values),max=Math.max(...values),span=max-min||1,y0=top+i*rowH+20,points=values.map((v,j)=>{const x=left+j*(w-left-right)/Math.max(1,s.length-1),y=y0+18-(v-min)/span*34;return {x,y}});svg+='<text class="series-name" x="12" y="'+(y0+2)+'">'+key+'</text><line x1="'+left+'" y1="'+(y0+20)+'" x2="'+(w-right)+'" y2="'+(y0+20)+'" stroke="#d8e1d8"/>';svg+='<polyline fill="none" stroke="'+colors[i]+'" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" points="'+points.map(p=>p.x+','+p.y).join(' ')+'"/>';points.forEach(p=>{svg+='<circle cx="'+p.x+'" cy="'+p.y+'" r="3" fill="#fff" stroke="'+colors[i]+'" stroke-width="2"/>'});svg+='<text class="last-value" x="'+(w-right+8)+'" y="'+(y0+4)+'" fill="'+colors[i]+'">'+display(values[values.length-1],key)+'</text>'});const step=Math.max(1,Math.ceil(s.length/8));s.forEach((x,i)=>{if(i%step===0||i===s.length-1){const px=left+i*(w-left-right)/Math.max(1,s.length-1);svg+='<text class="axis" text-anchor="middle" x="'+px+'" y="'+(h-8)+'">'+x.label+'</text>'}});document.querySelector('#chart').className='core-chart';document.querySelector('#chart').innerHTML=svg+'</svg>'}
function renderCore(){const s=seriesForRange(),last=s[s.length-1]?.value||{};document.querySelector('#kpis').innerHTML=[['净销售额',last['净销售额']],['客单价',last['客单价']],['老客占比',last['老客占比']],['费比',last['费比']]].map(x=>'<div class="kpi core-kpi"><span class="kpi-label">'+x[0]+' · '+(s[s.length-1]?.label||'--')+'</span><strong class="kpi-value">'+display(x[1],x[0])+'</strong></div>').join('');let table='<table><thead><tr><th>分类</th><th>指标</th>'+s.map(x=>'<th>'+x.label+'</th>').join('')+'</tr></thead><tbody>';P.sections.forEach(g=>g.metrics.forEach((metric,index)=>{const cls=focus.has(metric)?' focus-row':'';const label=focus.has(metric)?'<span class="focus-flag">重点</span>'+metric:metric;table+='<tr class="'+cls.trim()+'" data-section="'+g.name+'">'+(index===0?'<td class="section" rowspan="'+g.metrics.length+'">'+g.name+'</td>':'')+'<td class="metric">'+label+'</td>'+s.map(x=>'<td>'+display(x.value[metric],metric)+'</td>').join('')+'</tr>'}));document.querySelector('#table').innerHTML=table+'</tbody></table>';document.querySelector('#status').textContent=(granularity==='day'?'按日':granularity==='week'?'按周':'按月')+' · '+s.length+' 个时间段';renderCoreChart(s)}
renderView=renderCore;
'''
    # The original dashboard renderer is retained for the table fallback, but the final
    # renderer must own its calculations and run only after all functions are defined.
    time_script = time_script.replace("renderCalendar();renderView();", "renderCalendar();")
    core_logic = """const focus=new Set(['净销售额','客单价','老客占比','费比']);const raw=['支付金额','成功退款金额','商品访客数','支付买家数','支付件数','支付新买家数','支付老买家数','商品加购人数','点击量','花费','成交人数','总成交金额','总成交笔数','总购物车数'];const percent=new Set(['退款率','商品支付转化率','新客占比','老客占比','加购率','费比']);const display=(v,k)=>v==null?'--':(percent.has(k)?(v*100).toFixed(1)+'%':Number(v).toLocaleString('zh-CN',{maximumFractionDigits:1}));function finalize(v){v['净销售额']=Number(v['支付金额']||0)-Number(v['成功退款金额']||0);v['退款率']=v['支付金额']?v['成功退款金额']/v['支付金额']:null;v['商品支付转化率']=v['商品访客数']?v['支付买家数']/v['商品访客数']:null;v['客单价']=v['支付买家数']?v['支付金额']/v['支付买家数']:null;v['新客占比']=v['支付买家数']?v['支付新买家数']/v['支付买家数']:null;v['老客占比']=v['支付买家数']?v['支付老买家数']/v['支付买家数']:null;v['加购率']=v['商品访客数']?v['商品加购人数']/v['商品访客数']:null;v['费比']=v['净销售额']?v['花费']/v['净销售额']:null;v['投产']=v['花费']?v['总成交金额']/v['花费']:null;return v}function summary(s){const v={};raw.forEach(k=>v[k]=s.reduce((n,x)=>n+Number(x.value[k]||0),0));return finalize(v)}""" + core_logic
    core_logic = core_logic.replace("const s=seriesForRange(),last=s[s.length-1]?.value||{};", "const s=seriesForRange(),last=summary(s);")
    core_logic = core_logic.replace("function renderCore(){const s=seriesForRange(),last=summary(s);", "function renderCore(){document.querySelector('.panel-head').textContent='核心指标趋势';const s=seriesForRange(),last=summary(s),period=datesForRange();document.querySelector('#time-label').textContent=(period[0]||'--')+' ~ '+(period[period.length-1]||'--');")
    core_logic = core_logic.replace("function renderCore(){document.querySelector('.panel-head').textContent='核心指标趋势';", "function renderCoreBase(){document.querySelector('.panel-head').textContent='核心指标趋势';")
    core_logic = core_logic.replace("+(s[s.length-1]?.label||'--')", "+'统计区间'")
    link_css = """<style>
    .kpi.core-kpi{appearance:none;border:0;border-right:1px solid var(--line);font:inherit;text-align:left;cursor:pointer;transition:background .15s ease,box-shadow .15s ease}.kpi.core-kpi:last-child{border-right:0}.kpi.core-kpi:hover{background:#f4f8f3;box-shadow:inset 0 -3px 0 #7fa58a}.kpi.core-kpi:focus-visible{outline:2px solid #587a9b;outline-offset:-2px}.kpi.core-kpi[aria-pressed="true"]{box-shadow:inset 0 -3px 0 #2e6b3f,0 0 0 1px #2e6b3f}.core-chart .metric-series{transition:opacity .15s ease,stroke-width .15s ease}.core-chart .metric-series.is-muted{opacity:.2}.core-chart .metric-series.is-selected{opacity:1;stroke-width:3.5}.core-chart .series-name{cursor:pointer}.core-chart .series-name.is-muted{opacity:.42}.core-chart .series-name.is-selected{font-weight:800}
    </style>"""
    modern_css = """<style>
    :root{--paper:#f3f6f8;--line:#d6e0e6;--forest:#203746;--forest-dark:#17303e;--ink:#172630;--muted:#6f7f89;--mint:#e4f2ee;--mint-soft:#f3f8f6;--blue:#e8f0f7;--yellow:#fbf2d8;--accent:#157b7c;--accent-soft:#e5f3f2;--danger:#9d4a39}
    *{box-sizing:border-box}html{scroll-behavior:smooth}body{background:var(--paper);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif;font-size:13px;letter-spacing:0}header{background:linear-gradient(115deg,#1b303d 0%,#294858 100%);padding:18px clamp(20px,4vw,58px);border-bottom:1px solid #385767}.mast{max-width:1480px;display:grid;grid-template-columns:auto 1fr auto;align-items:center;gap:34px}.eyebrow{color:#a7c7cc;letter-spacing:.14em;font-size:10px}.title{font-size:27px;font-weight:700;letter-spacing:.02em}.app-nav{display:flex;align-self:stretch;align-items:stretch;gap:4px}.app-nav a{display:flex;align-items:center;padding:0 13px;color:#b8cbd2;text-decoration:none;font-weight:600;font-size:12px;border-bottom:2px solid transparent}.app-nav a:hover{color:#fff;border-bottom-color:#79aeb2}.app-nav a.active{color:#fff;border-bottom-color:#76c5be}.meta{max-width:420px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:#c8d8de;line-height:1.6}.shell{max-width:1480px;padding:18px clamp(20px,4vw,58px) 42px}.toolbar{background:#fff;border:1px solid var(--line);border-radius:10px;padding:14px 16px;margin-bottom:14px;box-shadow:0 3px 12px rgba(23,38,48,.05)}.time-line{margin-bottom:11px}.time-line strong{color:#2d7080!important;font-weight:600}.time-group{border-color:#d8e3e8;background:#f7fafb;border-radius:8px}.time-group button{color:#3b5361;border-radius:6px}.time-group button:hover{background:#eaf3f4;color:#176d72}.time-group button[aria-pressed="true"]{background:var(--accent-soft);color:#126a6b;box-shadow:inset 0 0 0 1px #9ccfcb}.time-divider{background:#d6e0e6}.status{color:#778892}.kpis{gap:12px;border:0;background:transparent;margin:0 0 14px}.kpi.core-kpi{border:1px solid var(--line);border-top:3px solid #69a7a1;border-radius:8px;background:#fff!important;box-shadow:0 3px 12px rgba(23,38,48,.05);min-height:88px}.kpi.core-kpi:nth-child(2){border-top-color:#6f91ae}.kpi.core-kpi:nth-child(3){border-top-color:#7aa15a}.kpi.core-kpi:nth-child(4){border-top-color:#b08a38}.kpi.core-kpi:hover{background:#fbfdfd!important}.kpi.core-kpi[aria-pressed="true"]{box-shadow:0 4px 15px rgba(21,123,124,.14),inset 0 -3px 0 var(--accent)}.kpi-label{color:#6a7b84;font-size:11px}.kpi-value{font-size:24px;color:#1a303c}.viz{gap:14px;margin-bottom:14px}.panel{border:1px solid var(--line);border-radius:10px;box-shadow:0 3px 12px rgba(23,38,48,.04);overflow:hidden}.panel-head{background:#fff;padding:13px 16px;border-bottom:1px solid #e1e8ec;color:#203746;font-size:14px}.core-chart{background:#fff;height:336px;padding:15px 18px}.core-chart .axis{fill:#74858d}.core-chart .series-name{fill:#354d59}.notes{background:#fff;padding:16px}.notes h2{color:#203746}.notes p{color:#72818a}.report-table{border:1px solid var(--line);border-radius:10px;box-shadow:0 3px 12px rgba(23,38,48,.04);overflow:auto}.report-table table{min-width:1020px}.report-table th{background:#203c4b;color:#f4f8f9;border-color:#365563;padding:11px 12px;font-size:12px}.report-table td{border-color:#dbe4e8;padding:9px 12px;background:#fff}.report-table td.metric{background:#f0f6f4;color:#243a43}.report-table tr[data-section="流量数据"] td.metric{background:#edf6f7}.report-table tr[data-section="付费数据"] td.metric{background:#fbf5de}.report-table td.section{background:#d8ece5;color:#2d5c57;border-color:#c9ded8}.report-table tr[data-section="流量数据"] td.section{background:#d8eff0;color:#2e666a}.report-table tr[data-section="付费数据"] td.section{background:#f7e8b0;color:#725d1e}.report-table tr.focus-row td.metric{background:#fff1eb;color:var(--danger)}.report-table tr.focus-row td:not(.metric):not(.section){background:#fffaf7}.foot{color:#83919a;padding-top:15px}@media(max-width:840px){header{padding:16px 18px}.mast{display:flex;flex-wrap:wrap;gap:12px}.app-nav{order:3;width:100%;height:34px}.app-nav a{padding:0 10px}.meta{max-width:100%;margin-left:auto}.toolbar{padding:12px}.kpis{gap:8px}.kpi.core-kpi{min-height:78px}.kpi-value{font-size:21px}.core-chart{height:360px;padding:12px;overflow-x:auto}.core-chart svg{min-width:760px}.report-table table{min-width:960px}}
    </style>"""
    link_logic = r'''let selectedMetric='净销售额';const focusKeys=['净销售额','客单价','老客占比','费比'];const baseRenderCore=renderCoreBase;function selectMetric(key){selectedMetric=key;document.querySelectorAll('#kpis [data-metric]').forEach(card=>card.setAttribute('aria-pressed',String(card.dataset.metric===key)));document.querySelectorAll('#chart [data-metric]').forEach(node=>{node.classList.toggle('is-selected',node.dataset.metric===key);node.classList.toggle('is-muted',node.dataset.metric!==key)})}function enhanceMetricLinks(){const cards=[...document.querySelectorAll('#kpis .core-kpi')];cards.forEach((card,index)=>{const key=focusKeys[index];const next=document.createElement('button');next.type='button';next.className=card.className;next.dataset.metric=key;next.setAttribute('aria-pressed',String(key===selectedMetric));next.innerHTML=card.innerHTML;card.replaceWith(next);next.addEventListener('click',()=>selectMetric(key));next.addEventListener('keydown',event=>{if(event.key==='Enter'||event.key===' '){event.preventDefault();selectMetric(key)}})});const labels=[...document.querySelectorAll('#chart .series-name')];labels.forEach((label,index)=>{const key=focusKeys[index];label.dataset.metric=key;label.addEventListener('click',()=>selectMetric(key));let node=label.nextElementSibling;while(node&&!node.classList.contains('axis')){node.dataset.metric=key;node.classList.add('metric-series');node.addEventListener('click',()=>selectMetric(key));node=node.nextElementSibling}});selectMetric(selectedMetric)}function renderCore(){baseRenderCore();enhanceMetricLinks()}
renderView=renderCore;renderCalendar();renderView();
'''
    core_logic = core_logic.replace("renderView=renderCore;", link_logic)
    time_script = time_script.replace("\n})();\n</script>", "\n" + core_logic + "})();\n</script>")
    analysis_css = """<style>
    .app-nav button{appearance:none;background:transparent;border:0;border-bottom:2px solid transparent;color:#b8cbd2;padding:0 13px;font:600 12px inherit;cursor:pointer}.app-nav button:hover{color:#fff;border-bottom-color:#79aeb2}.app-nav button.active{color:#fff;border-bottom-color:#76c5be}.analysis-view{padding-top:2px}.analysis-heading{display:flex;align-items:flex-end;justify-content:space-between;gap:16px;margin:2px 0 14px}.analysis-heading h2{margin:0;color:#203746;font-size:21px}.analysis-heading p{margin:5px 0 0;color:#72818a;font-size:12px}.analysis-heading>span{color:#778892;font-size:12px}.analysis-kpis{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-bottom:14px}.analysis-kpi{background:#fff;border:1px solid var(--line);border-top:3px solid #6ea7a2;border-radius:8px;padding:13px 15px;min-height:82px}.analysis-kpi:nth-child(2){border-top-color:#6f91ae}.analysis-kpi:nth-child(3){border-top-color:#b08a38}.analysis-kpi:nth-child(4){border-top-color:#7aa15a}.analysis-kpi span{display:block;color:#6a7b84;font-size:11px}.analysis-kpi strong{display:block;margin-top:7px;color:#1a303c;font-size:22px}.analysis-grid{display:grid;grid-template-columns:minmax(250px,.42fr) minmax(0,1.58fr);gap:14px}.analysis-panel{background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:0 3px 12px rgba(23,38,48,.04);overflow:hidden}.analysis-panel h3{margin:0;padding:13px 15px;border-bottom:1px solid #e1e8ec;color:#203746;font-size:13px}.analysis-panel h3 small{margin-left:8px;color:#829099;font-size:10px;font-weight:400}.rank-list{padding:8px 14px 12px}.rank-item{display:grid;grid-template-columns:24px 1fr auto;gap:9px;align-items:center;padding:11px 0;border-bottom:1px solid #edf1f2}.rank-item:last-child{border-bottom:0}.rank-num{width:22px;height:22px;border-radius:50%;display:grid;place-items:center;background:#edf5f1;color:#2d665b;font-size:11px;font-weight:700}.rank-main{min-width:0}.rank-name{display:block;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#27414b;font-weight:600}.rank-sub{display:block;margin-top:3px;color:#829099;font-size:11px}.rank-value{color:#245f67;font-weight:700;white-space:nowrap}.analysis-table{min-width:0}.analysis-table-wide{width:100%}.analysis-table>div{overflow:auto;max-height:620px}.analysis-table table{width:100%;min-width:1480px;border-collapse:collapse}.analysis-table th,.analysis-table td{padding:9px 10px;border-bottom:1px solid #e7edef;border-right:1px solid #eef2f3;text-align:right;white-space:nowrap}.analysis-table thead{position:sticky;top:0;z-index:5}.analysis-table th{background:#203c4b;color:#f4f8f9;font-size:11px}.analysis-table th:first-child,.analysis-table td:first-child{text-align:left;position:sticky;left:0;z-index:2}.analysis-table th:first-child{z-index:6}.analysis-table td:first-child{background:#fff;font-weight:600;color:#29434d;max-width:280px;overflow:hidden;text-overflow:ellipsis}.analysis-table td{font-size:12px}.analysis-table tbody tr:hover td{background:#f5faf8}.analysis-table tbody tr:hover td:first-child{background:#eef6f3}.table-head{display:flex;align-items:center;gap:5px;justify-content:flex-end}.table-head.left{justify-content:flex-start}.sort-btn{appearance:none;border:0;background:transparent;color:inherit;padding:0 2px;cursor:pointer;font-size:10px;opacity:.7}.sort-btn:hover,.sort-btn.active{opacity:1;color:#8ed4cc}.filter-row th{background:#294958;padding:6px}.filter-input{width:100%;min-width:78px;border:1px solid #54707c;border-radius:4px;background:#fff;color:#203746;padding:5px 7px;font:11px inherit}.filter-input:focus{outline:2px solid #87c9c2;outline-offset:1px}.report-table .table-head{justify-content:center}.report-table .sort-btn{color:#d8e9ec}.report-table .filter-row th{position:sticky;top:37px;z-index:4}.report-table .filter-input{min-width:92px}.analysis-empty{padding:30px 16px;color:#7e8e95;text-align:center}.analysis-view[hidden]{display:none!important}@media(max-width:840px){.app-nav button{padding:0 10px}.analysis-heading{align-items:flex-start;flex-direction:column}.analysis-kpis{grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}.analysis-kpi{min-height:74px;padding:11px 12px}.analysis-kpi strong{font-size:19px}.analysis-grid{grid-template-columns:1fr}.analysis-panel{overflow:hidden}.analysis-table table{min-width:1420px}.report-table .filter-row th{top:35px}}
    </style>"""
    analysis_css += "<style>.filter-row{display:none!important}.analysis-active>#core,.analysis-active>#kpis,.analysis-active>#warn,.analysis-active>#table{display:none!important}.analysis-heading-tools{display:flex;align-items:center;gap:12px}.analysis-search{width:min(320px,42vw);border:1px solid #cbd9df;border-radius:6px;background:#fff;color:#203746;padding:9px 11px;font:12px inherit}.analysis-search:focus{outline:2px solid #87c9c2;outline-offset:1px}.analysis-rank-wide{margin-bottom:14px}.analysis-rank-wide .rank-list{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:0 22px;padding:6px 16px 10px}.analysis-rank-wide .rank-item{min-width:0}.analysis-detail-wide{width:100%}@media(max-width:840px){.analysis-heading-tools{width:100%;align-items:flex-start;flex-direction:column}.analysis-search{width:100%}.analysis-rank-wide .rank-list{grid-template-columns:repeat(2,minmax(0,1fr));gap:0 12px}}</style>"
    html = html.replace("</style>", css + focus_css + core_css + time_css + link_css + modern_css + analysis_css + "</style>", 1)
    analysis_script = r'''<script>
(function(){
  const payload=window.__TMALL_ANALYSIS__||P.analysis||{products:[],promotions:[]};
  const productRows=Array.isArray(payload.products)?payload.products:[], promotionRows=Array.isArray(payload.promotions)?payload.promotions:[];
  const key={date:'日期',pid:'商品ID',pname:'商品名称',sceneId:'场景ID',sceneName:'场景名字',planId:'计划ID',planName:'计划名字',sales:'支付金额',refund:'成功退款金额',visitors:'商品访客数',buyers:'支付买家数',orders:'支付件数',adders:'商品加购人数',impressions:'展现量',clicks:'点击量',spend:'花费',gross:'总成交金额',roi:'投入产出比',cart:'总购物车数',people:'成交人数'};
  const num=(v)=>Number.isFinite(Number(v))?Number(v):0, text=(v)=>String(v??'').trim()||'未命名';
  const money=(v)=>'¥'+num(v).toLocaleString('zh-CN',{maximumFractionDigits:0}), integer=(v)=>num(v).toLocaleString('zh-CN',{maximumFractionDigits:0}), pct=(v)=>num(v*100).toFixed(1)+'%';
  function activeDates(){const pressed=document.querySelector('.time-group button[aria-pressed="true"][data-range]');const mode=pressed?.dataset.range||'7';if(mode==='yesterday')return P.dates.slice(-1);if(mode==='7')return P.dates.slice(-7);if(mode==='30')return P.dates.slice(-30);if(mode==='60')return P.dates.slice(-60);const label=document.querySelector('#time-label')?.textContent||'';const parts=label.match(/\d{4}-\d{2}-\d{2}/g);return parts&&parts.length>1?P.dates.filter(d=>d>=parts[0]&&d<=parts[1]):P.dates.slice(-7)}
  function selectedRows(rows){const set=new Set(activeDates());return rows.filter(r=>!r[key.date]||set.has(r[key.date]))}
  function productAggregate(){const groups=new Map();selectedRows(productRows).forEach(r=>{const id=text(r[key.pid]);if(!id)return;const g=groups.get(id)||{id,name:text(r[key.pname]),visitors:0,buyers:0,orders:0,adders:0,sales:0,refund:0};g.visitors+=num(r[key.visitors]);g.buyers+=num(r[key.buyers]);g.orders+=num(r[key.orders]);g.adders+=num(r[key.adders]);g.sales+=num(r[key.sales]);g.refund+=num(r[key.refund]);if(g.name==='未命名')g.name=text(r[key.pname]);groups.set(id,g)});return [...groups.values()].map(g=>({...g,net:g.sales-g.refund,conversion:g.visitors?g.buyers/g.visitors:0,atc:g.visitors?g.adders/g.visitors:0,refundRate:g.sales?g.refund/g.sales:0,aov:g.buyers?g.sales/g.buyers:0})).sort((a,b)=>b.net-a.net)}
  function promotionAggregate(){const groups=new Map();selectedRows(promotionRows).forEach(r=>{const id=text(r[key.planId])!=='未命名'?text(r[key.planId]):text(r[key.sceneId]);if(!id)return;const name=text(r[key.planName])!=='未命名'?text(r[key.planName]):text(r[key.sceneName]);const g=groups.get(id)||{id,name,impressions:0,clicks:0,spend:0,gross:0,roi:0,cart:0,people:0};g.impressions+=num(r[key.impressions]);g.clicks+=num(r[key.clicks]);g.spend+=num(r[key.spend]);g.gross+=num(r[key.gross]);g.cart+=num(r[key.cart]);g.people+=num(r[key.people]);groups.set(id,g)});return [...groups.values()].map(g=>({...g,ctr:g.impressions?g.clicks/g.impressions:0,cpc:g.clicks?g.spend/g.clicks:0,roi:g.spend?g.gross/g.spend:0,cartRate:g.clicks?g.cart/g.clicks:0})).sort((a,b)=>b.spend-a.spend)}
  function kpi(items,label,value){return '<div class="analysis-kpi"><span>'+label+'</span><strong>'+value+'</strong></div>'}
  function renderProducts(){const rows=productAggregate(),total=rows.reduce((a,r)=>{a.visitors+=r.visitors;a.buyers+=r.buyers;a.sales+=r.sales;a.refund+=r.refund;a.adders+=r.adders;return a},{visitors:0,buyers:0,sales:0,refund:0,adders:0});document.querySelector('#product-count').textContent=rows.length+' 个商品';document.querySelector('#product-kpis').innerHTML=[kpi(rows,'商品数',integer(rows.length)),kpi(rows,'净销售额',money(total.sales-total.refund)),kpi(rows,'支付买家数',integer(total.buyers)),kpi(rows,'退款率',pct(total.sales?total.refund/total.sales:0))].join('');document.querySelector('#product-rank').innerHTML=rows.slice(0,6).map((r,i)=>'<div class="rank-item"><span class="rank-num">'+(i+1)+'</span><div class="rank-main"><span class="rank-name">'+r.name+'</span><span class="rank-sub">'+r.id+' · '+integer(r.buyers)+' 买家</span></div><span class="rank-value">'+money(r.net)+'</span></div>').join('')||'<div class="analysis-empty">当前时间范围暂无商品数据</div>';document.querySelector('#product-table').innerHTML=rows.length?'<table><thead><tr><th>商品</th><th>访客数</th><th>支付买家</th><th>净销售额</th><th>客单价</th><th>转化率</th><th>加购率</th><th>退款率</th></tr></thead><tbody>'+rows.map(r=>'<tr><td>'+r.name+'<br><small>'+r.id+'</small></td><td>'+integer(r.visitors)+'</td><td>'+integer(r.buyers)+'</td><td>'+money(r.net)+'</td><td>'+money(r.aov)+'</td><td>'+pct(r.conversion)+'</td><td>'+pct(r.atc)+'</td><td>'+pct(r.refundRate)+'</td></tr>').join('')+'</tbody></table>':'<div class="analysis-empty">当前时间范围暂无商品数据</div>'}
  function renderPromotions(){const rows=promotionAggregate(),total=rows.reduce((a,r)=>{a.impressions+=r.impressions;a.clicks+=r.clicks;a.spend+=r.spend;a.gross+=r.gross;a.cart+=r.cart;a.people+=r.people;return a},{impressions:0,clicks:0,spend:0,gross:0,cart:0,people:0});document.querySelector('#promotion-count').textContent=rows.length+' 个计划 / 场景';document.querySelector('#promotion-kpis').innerHTML=[kpi(rows,'推广花费',money(total.spend)),kpi(rows,'总成交金额',money(total.gross)),kpi(rows,'整体投产',total.spend?(total.gross/total.spend).toFixed(2):'--'),kpi(rows,'点击量',integer(total.clicks))].join('');document.querySelector('#promotion-rank').innerHTML=rows.slice(0,6).map((r,i)=>'<div class="rank-item"><span class="rank-num">'+(i+1)+'</span><div class="rank-main"><span class="rank-name">'+r.name+'</span><span class="rank-sub">'+r.id+' · '+integer(r.clicks)+' 点击</span></div><span class="rank-value">'+money(r.spend)+'</span></div>').join('')||'<div class="analysis-empty">当前时间范围暂无推广数据</div>';document.querySelector('#promotion-table').innerHTML=rows.length?'<table><thead><tr><th>计划 / 场景</th><th>展现量</th><th>点击量</th><th>花费</th><th>点击率</th><th>CPC</th><th>成交金额</th><th>ROI</th><th>加购率</th></tr></thead><tbody>'+rows.map(r=>'<tr><td>'+r.name+'<br><small>'+r.id+'</small></td><td>'+integer(r.impressions)+'</td><td>'+integer(r.clicks)+'</td><td>'+money(r.spend)+'</td><td>'+pct(r.ctr)+'</td><td>'+money(r.cpc)+'</td><td>'+money(r.gross)+'</td><td>'+r.roi.toFixed(2)+'</td><td>'+pct(r.cartRate)+'</td></tr>').join('')+'</tbody></table>':'<div class="analysis-empty">当前时间范围暂无推广数据</div>'}
  function productAggregateFull(){const groups=new Map();selectedRows(productRows).forEach(r=>{const id=text(r[key.pid]);if(!id)return;const g=groups.get(id)||{id,name:text(r[key.pname]),layer:text(r['分层']),visitors:0,views:0,stay:0,bounce:0,favorites:0,addOrders:0,buyers:0,orders:0,adders:0,sales:0,refund:0,promoIpv:0,promoSpend:0,promoRoi:0,collectRate:0,repurchaseRate:0,organicIpv:0,searchIpv:0,recommendIpv:0};g.visitors+=num(r[key.visitors]);g.views+=num(r['商品浏览量']);g.stay+=num(r['平均停留时长']);g.bounce+=num(r['商品详情页跳出率']);g.favorites+=num(r['商品收藏人数']);g.addOrders+=num(r['商品加购件数']);g.adders+=num(r[key.adders]);g.buyers+=num(r[key.buyers]);g.orders+=num(r[key.orders]);g.sales+=num(r[key.sales]);g.refund+=num(r[key.refund]);g.promoIpv+=num(r['营销推广IPV']);g.promoSpend+=num(r['营销推广消耗']);g.promoRoi+=num(r['营销推广ROI']);g.collectRate+=num(r['收加率']);g.repurchaseRate+=num(r['复购率']);g.organicIpv+=num(r['非推广IPV']);g.searchIpv+=num(r['搜索IPV']);g.recommendIpv+=num(r['推荐IPV']);groups.set(id,g)});return [...groups.values()].map(g=>{const net=g.sales-g.refund;return {...g,net,productFeeRatio:net?g.promoSpend/net:null,conversion:g.visitors?g.buyers/g.visitors:0,atc:g.visitors?g.adders/g.visitors:0,refundRate:g.sales?g.refund/g.sales:0,aov:g.buyers?g.sales/g.buyers:0}}).sort((a,b)=>b.net-a.net)}
  function promotionAggregateFull(){const groups=new Map();selectedRows(promotionRows).forEach(r=>{const id=text(r[key.planId])!=='未命名'?text(r[key.planId]):text(r[key.sceneId]);if(!id)return;const name=text(r[key.planName])!=='未命名'?text(r[key.planName]):text(r[key.sceneName]);const g=groups.get(id)||{id,name,impressions:0,clicks:0,spend:0,gross:0,roi:0,cart:0,people:0,orders:0,guided:0,newPeople:0};g.impressions+=num(r[key.impressions]);g.clicks+=num(r[key.clicks]);g.spend+=num(r[key.spend]);g.gross+=num(r[key.gross]);g.cart+=num(r[key.cart]);g.people+=num(r[key.people]);g.orders+=num(r['总成交笔数']);g.guided+=num(r['引导访问人数']);g.newPeople+=num(r['成交新客数']);groups.set(id,g)});return [...groups.values()].map(g=>({...g,ctr:g.impressions?g.clicks/g.impressions:0,cpc:g.clicks?g.spend/g.clicks:0,roi:g.spend?g.gross/g.spend:0,cartRate:g.clicks?g.cart/g.clicks:0,newRate:g.people?g.newPeople/g.people:0})).sort((a,b)=>b.spend-a.spend)}
  function renderProductsFull(){const rows=productAggregateFull(),total=rows.reduce((a,r)=>{a.visitors+=r.visitors;a.buyers+=r.buyers;a.sales+=r.sales;a.refund+=r.refund;a.adders+=r.adders;a.orders+=r.orders;return a},{visitors:0,buyers:0,sales:0,refund:0,adders:0,orders:0});document.querySelector('#product-count').textContent=rows.length+' 个商品';document.querySelector('#product-kpis').innerHTML=[kpi(rows,'商品数',integer(rows.length)),kpi(rows,'净销售额',money(total.sales-total.refund)),kpi(rows,'支付买家数',integer(total.buyers)),kpi(rows,'退款率',pct(total.sales?total.refund/total.sales:0))].join('');document.querySelector('#product-rank').innerHTML=rows.slice(0,8).map((r,i)=>'<div class="rank-item"><span class="rank-num">'+(i+1)+'</span><div class="rank-main"><span class="rank-name">'+r.name+'</span><span class="rank-sub">'+r.id+' · '+integer(r.buyers)+' 买家 · '+pct(r.conversion)+' 转化</span></div><span class="rank-value">'+money(r.net)+'</span></div>').join('')||'<div class="analysis-empty">当前时间范围暂无商品数据</div>';const cols=[['商品','name'],['商品ID','id'],['分层','layer'],['访客数','visitors'],['浏览量','views'],['平均停留','stay'],['跳出率','bounce'],['收藏人数','favorites'],['加购件数','addOrders'],['加购人数','adders'],['支付买家','buyers'],['支付件数','orders'],['支付金额','sales'],['成功退款','refund'],['净销售额','net'],['产品费比','productFeeRatio'],['客单价','aov'],['支付转化率','conversion'],['加购率','atc'],['退款率','refundRate'],['推广IPV','promoIpv'],['推广消耗','promoSpend'],['推广ROI','promoRoi'],['收加率','collectRate'],['复购率','repurchaseRate'],['非推广IPV','organicIpv'],['搜索IPV','searchIpv'],['推荐IPV','recommendIpv']];const value=(r,k)=>({name:r.name,id:r.id,layer:r.layer||'--',visitors:integer(r.visitors),views:integer(r.views),stay:num(r.stay).toFixed(1)+'s',bounce:pct(num(r.bounce)/100),favorites:integer(r.favorites),addOrders:integer(r.addOrders),adders:integer(r.adders),buyers:integer(r.buyers),orders:integer(r.orders),sales:money(r.sales),refund:money(r.refund),net:money(r.net),productFeeRatio:pct(r.productFeeRatio),aov:money(r.aov),conversion:pct(r.conversion),atc:pct(r.atc),refundRate:pct(r.refundRate),promoIpv:integer(r.promoIpv),promoSpend:money(r.promoSpend),promoRoi:num(r.promoRoi).toFixed(2),collectRate:pct(num(r.collectRate)/100),repurchaseRate:pct(num(r.repurchaseRate)/100),organicIpv:integer(r.organicIpv),searchIpv:integer(r.searchIpv),recommendIpv:integer(r.recommendIpv)}[k]);document.querySelector('#product-table').innerHTML='<table><thead><tr>'+cols.map(c=>'<th data-filter-key="'+c[1]+'">'+c[0]+'</th>').join('')+'</tr></thead><tbody>'+rows.map(r=>'<tr>'+cols.map(c=>'<td>'+value(r,c[1])+'</td>').join('')+'</tr>').join('')+'</tbody></table>'}
  function renderPromotionsFull(){const rows=promotionAggregate(),total=rows.reduce((a,r)=>{a.impressions+=r.impressions;a.clicks+=r.clicks;a.spend+=r.spend;a.gross+=r.gross;a.cart+=r.cart;a.people+=r.people;return a},{impressions:0,clicks:0,spend:0,gross:0,cart:0,people:0});document.querySelector('#promotion-count').textContent=rows.length+' 个计划 / 场景';document.querySelector('#promotion-kpis').innerHTML=[kpi(rows,'推广花费',money(total.spend)),kpi(rows,'总成交金额',money(total.gross)),kpi(rows,'整体投产',total.spend?(total.gross/total.spend).toFixed(2):'--'),kpi(rows,'点击量',integer(total.clicks))].join('');const cols=[['计划 / 场景','name'],['计划ID','id'],['展现量','impressions'],['点击量','clicks'],['花费','spend'],['点击率','ctr'],['CPC','cpc'],['成交金额','gross'],['ROI','roi'],['成交笔数','orders'],['成交人数','people'],['购物车数','cart'],['加购率','cartRate'],['引导访问人数','guided'],['成交新客数','newPeople'],['成交新客占比','newRate']];const value=(r,k)=>({name:r.name,id:r.id,impressions:integer(r.impressions),clicks:integer(r.clicks),spend:money(r.spend),ctr:pct(r.ctr),cpc:money(r.cpc),gross:money(r.gross),roi:num(r.roi).toFixed(2),orders:integer(r.orders),people:integer(r.people),cart:integer(r.cart),cartRate:pct(r.cartRate),guided:integer(r.guided),newPeople:integer(r.newPeople),newRate:pct(r.newRate)}[k]);document.querySelector('#promotion-table').innerHTML='<table><thead><tr>'+cols.map(c=>'<th data-filter-key="'+c[1]+'">'+c[0]+'</th>').join('')+'</tr></thead><tbody>'+rows.map(r=>'<tr>'+cols.map(c=>'<td>'+value(r,c[1])+'</td>').join('')+'</tr>').join('')+'</tbody></table>'}
  function tableEnhance(table){if(!table||table.dataset.enhanced==='1')return;const head=table.querySelector('thead tr');const body=table.querySelector('tbody');if(!head||!body)return;const headers=[...head.children];const models=[...body.rows].map(row=>({cells:[...row.cells].map(c=>c.innerHTML),section:row.dataset.section||row.querySelector('.section')?.textContent||''}));table.dataset.enhanced='1';headers.forEach((th,i)=>{const label=th.textContent.trim();th.innerHTML='<span class="table-head'+(i<2?' left':'')+'">'+label+' <button type="button" class="sort-btn" data-sort-col="'+i+'" aria-label="按'+label+'排序">↕</button></span>'});const filter=document.createElement('tr');filter.className='filter-row';headers.forEach((th,i)=>{const cell=document.createElement('th');if(i>=0){cell.innerHTML='<input class="filter-input" type="search" placeholder="筛选" aria-label="筛选'+th.textContent.trim()+'" data-filter-col="'+i+'">'}filter.appendChild(cell)});table.querySelector('thead').appendChild(filter);let sortCol=-1,sortDir=1;function paint(){const terms=[...table.querySelectorAll('.filter-input')].map(x=>x.value.trim().toLowerCase());let visible=models.filter(m=>m.cells.every((v,i)=>!terms[i]||String(v).replace(/<[^>]+>/g,'').toLowerCase().includes(terms[i])));if(sortCol>=0)visible.sort((a,b)=>String(a.cells[sortCol]).replace(/<[^>]+>/g,'').localeCompare(String(b.cells[sortCol]).replace(/<[^>]+>/g,''),'zh-CN',{numeric:true})*sortDir);body.innerHTML=visible.map(m=>'<tr>'+m.cells.map((v,i)=>'<td>'+v+'</td>').join('')+'</tr>').join('');table.querySelectorAll('.sort-btn').forEach(b=>b.classList.toggle('active',Number(b.dataset.sortCol)===sortCol))}table.querySelectorAll('.filter-input').forEach(i=>i.addEventListener('input',paint));table.querySelectorAll('.sort-btn').forEach(b=>b.addEventListener('click',()=>{const c=Number(b.dataset.sortCol);sortDir=sortCol===c?-sortDir:1;sortCol=c;paint()}))}
  function normalizeOverview(table){if(!table||table.dataset.normalized==='1')return;let section='';[...table.tBodies[0].rows].forEach(row=>{const cell=row.querySelector('.section');if(cell){section=cell.textContent.trim();cell.removeAttribute('rowspan')}else if(section){const td=document.createElement('td');td.className='section';td.textContent=section;row.insertBefore(td,row.firstChild)}});table.dataset.normalized='1'}
  function enhanceTables(){document.querySelectorAll('#product-table table,#promotion-table table').forEach(table=>{tableEnhance(table);table.querySelector('.filter-row')?.remove()})}
  renderPromotionsFull=function(){const rows=promotionAggregateFull(),total=rows.reduce((a,r)=>{a.impressions+=r.impressions;a.clicks+=r.clicks;a.spend+=r.spend;a.gross+=r.gross;a.cart+=r.cart;a.people+=r.people;return a},{impressions:0,clicks:0,spend:0,gross:0,cart:0,people:0});document.querySelector('#promotion-count').textContent=rows.length+' 个计划 / 场景';document.querySelector('#promotion-kpis').innerHTML=[kpi(rows,'推广花费',money(total.spend)),kpi(rows,'总成交金额',money(total.gross)),kpi(rows,'整体投产',total.spend?(total.gross/total.spend).toFixed(2):'--'),kpi(rows,'点击量',integer(total.clicks))].join('');const cols=[['计划 / 场景','name'],['计划ID','id'],['展现量','impressions'],['点击量','clicks'],['花费','spend'],['点击率','ctr'],['CPC','cpc'],['成交金额','gross'],['ROI','roi'],['成交笔数','orders'],['成交人数','people'],['购物车数','cart'],['加购率','cartRate'],['引导访问人数','guided'],['成交新客数','newPeople'],['成交新客占比','newRate']];const value=(r,k)=>({name:r.name,id:r.id,impressions:integer(r.impressions),clicks:integer(r.clicks),spend:money(r.spend),ctr:pct(r.ctr),cpc:money(r.cpc),gross:money(r.gross),roi:num(r.roi).toFixed(2),orders:integer(r.orders),people:integer(r.people),cart:integer(r.cart),cartRate:pct(r.cartRate),guided:integer(r.guided),newPeople:integer(r.newPeople),newRate:pct(r.newRate)}[k]);document.querySelector('#promotion-table').innerHTML='<table><thead><tr>'+cols.map(c=>'<th>'+c[0]+'</th>').join('')+'</tr></thead><tbody>'+rows.map(r=>'<tr>'+cols.map(c=>'<td>'+value(r,c[1])+'</td>').join('')+'</tr>').join('')+'</tbody></table>'};function bindAnalysisSearches(){[['#product-search','#product-table'],['#promotion-search','#promotion-table']].forEach(([inputSel,tableSel])=>{const input=document.querySelector(inputSel),table=document.querySelector(tableSel);if(!input||!table)return;const fresh=input.cloneNode(true);input.replaceWith(fresh);fresh.addEventListener('input',()=>{const term=fresh.value.trim().toLowerCase();table.querySelectorAll('tbody tr').forEach(row=>{const text=[row.cells[0]?.textContent,row.cells[1]?.textContent].join(' ').toLowerCase();row.hidden=Boolean(term&&!text.includes(term))})})})}function renderAnalysis(){renderProductsFull();renderPromotionsFull();setTimeout(()=>{enhanceTables();bindAnalysisSearches()},0)}
  function switchView(view){document.querySelectorAll('[data-view]').forEach(b=>{const on=b.dataset.view===view;b.classList.toggle('active',on);b.setAttribute('aria-pressed',String(on))});const shell=document.querySelector('.shell');if(shell)shell.classList.toggle('analysis-active',view!=='overview');['#core','#kpis','#warn','#table'].forEach(sel=>{const node=document.querySelector(sel);if(node)node.hidden=view!=='overview'});document.querySelector('#product-analysis').hidden=view!=='product';document.querySelector('#promotion-analysis').hidden=view!=='promotion';if(view!=='overview')renderAnalysis()}
  document.querySelectorAll('[data-view]').forEach(b=>b.addEventListener('click',()=>switchView(b.dataset.view)));document.querySelectorAll('.time-group [data-range],.time-group [data-gran]').forEach(b=>b.addEventListener('click',()=>setTimeout(renderAnalysis,0)));const tableObserver=new MutationObserver(()=>setTimeout(enhanceTables,0));const tableRoot=document.querySelector('#table');if(tableRoot)tableObserver.observe(tableRoot,{childList:true});window.__tmallSwitchView=switchView;renderAnalysis();setTimeout(enhanceTables,0);
})();
</script>'''
    html = html.replace("</body>", focus_script + time_script + analysis_script + "</body>", 1)
    path.write_text(html, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--source-file", type=Path)
    args = parser.parse_args()
    analysis: dict[str, list[dict[str, Any]]] = {"products": [], "promotions": []}
    if args.source_file:
        source_label = args.source_file.name
        dates, matrix, warnings = collect_metrics(args.source_file)
    elif any(p.suffix.lower() in {".xls", ".zip"} for p in args.source_dir.iterdir()):
        dates, matrix, warnings, source_label = collect_metrics_directory(args.source_dir)
        analysis = collect_analysis_directory(args.source_dir)
    else:
        source = newest_workbook(args.source_dir, args.output_dir)
        source_label = source.name
        dates, matrix, warnings = collect_metrics(source)
    if not dates:
        raise RuntimeError("No dated rows found in 商品数据 or 推广数据")
    stamp = datetime.now().strftime("%Y%m%d")
    write_excel(args.output_dir / f"经营数据汇总_{stamp}.xlsx", dates, matrix, source_label)
    dashboard_dir = args.output_dir / "dashboard"
    write_dashboard_with_periods(dashboard_dir, dates, matrix, warnings, source_label, analysis)
    refine_dashboard_file(dashboard_dir / "index.html")
    print(json.dumps({"source": source_label, "dates": dates, "warnings": warnings, "excel": str(args.output_dir / f'经营数据汇总_{stamp}.xlsx'), "dashboard": str(args.output_dir / 'dashboard' / 'index.html')}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
