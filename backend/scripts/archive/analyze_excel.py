import openpyxl
import sys

files = [
    r"E:\桌面\0425\海贝海-5月规划-5.3.xlsx",
    r"E:\桌面\0409\海贝海-数据分析表-月.xlsx",
]

for f in files:
    print(f"\n{'='*80}")
    print(f"文件: {f}")
    print(f"{'='*80}")
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"工作表: {wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"\n--- {ws.title} ---")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
        print("列头:")
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val is not None:
                print(f"  列{c}: {val}")
        print("\n示例数据(前3行):")
        for r in range(2, min(5, ws.max_row + 1)):
            row_data = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_data.append(str(val)[:30])
            if row_data:
                print(f"  行{r}: {' | '.join(row_data[:8])}")
