import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
import sqlite3

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'dashboard.db')
PLANNING_FILE = r"E:\桌面\0425\海贝海-5月规划-5.3.xlsx"
MONTHLY_FILE = r"E:\桌面\0409\海贝海-数据分析表-月.xlsx"

def safe_float(val):
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def safe_int(val):
    if val is None or val == '':
        return 0
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return 0

def normalize_date(val):
    if val is None:
        return ''
    s = str(val).strip()
    if '至' in s:
        return s.split('至')[0].strip()
    return s

def extract_pid(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.replace('.','').isdigit():
        return str(int(float(s)))
    return s

def map_row_planning(ws, row, col_start=1):
    """Map row for monthly_planning table."""
    def c(offset):
        return ws.cell(row=row, column=col_start + offset - 1).value
    return {
        'product_id': extract_pid(c(1)),
        'product_title': c(2),
        'image_url': c(3),
        'category': c(4),
        'list_date': c(5),
        'payment_amount': safe_float(c(6)),
        'refund_amount': safe_float(c(7)),
        'net_sales': safe_float(c(8)),
        'ad_spend': safe_float(c(9)),
        'ad_roi': safe_float(c(10)),
        'ad_ratio': safe_float(c(11)),
        'visitors': safe_float(c(12)),
        'uv_value': safe_float(c(13)),
        'payment_conversion': safe_float(c(14)),
        'refund_rate': safe_float(c(15)),
        'cart_rate': safe_float(c(16)),
        'fav_rate': safe_float(c(17)),
        'bounce_rate': safe_float(c(18)),
        'avg_stay_duration': safe_float(c(19)),
        'search_conversion': safe_float(c(20)),
        'score': safe_float(c(21)),
        'page_views': safe_float(c(22)),
        'payment_buyers': safe_float(c(23)),
        'aov': safe_float(c(24)),
        'payment_qty': safe_float(c(25)),
        'cart_qty': safe_float(c(26)),
        'fav_users': safe_float(c(27)),
        'search_buyers': safe_float(c(28)),
        'search_ratio': safe_float(c(29)),
        'search_uv_value': safe_float(c(30)),
        'total_orders': safe_float(c(31)),
        'store_favs': safe_float(c(32)),
        'total_favs': safe_float(c(33)),
        'fav_cost': safe_float(c(34)),
        'fav_cart_qty': safe_float(c(35)),
        'order_conversion': safe_float(c(36)),
        'ctr': safe_float(c(37)),
        'total_ad_spend': safe_float(c(38)),
        'impressions': safe_float(c(39)),
        'clicks': safe_float(c(40)),
        'cart_items': safe_float(c(41)),
        'keyword_ad_spend': safe_float(c(42)),
        'keyword_ad_sales': safe_float(c(43)),
        'keyword_ad_roi': safe_float(c(44)),
        'keyword_ad_conversion': safe_float(c(45)),
        'keyword_impressions': safe_float(c(46)),
        'keyword_ctr': safe_float(c(47)),
        'keyword_visitors': safe_float(c(48)),
        'keyword_visitor_ratio': safe_float(c(49)),
        'keyword_ppc': safe_float(c(50)),
        'keyword_conversion_cost': safe_float(c(51)),
        'keyword_cart_qty': safe_float(c(52)),
        'keyword_cart_rate': safe_float(c(53)),
        'keyword_cart_cost': safe_float(c(54)),
        'audience_ad_spend': safe_float(c(55)),
        'audience_ad_sales': safe_float(c(56)),
        'audience_ad_roi': safe_float(c(57)),
        'audience_ad_conversion': safe_float(c(58)),
        'audience_impressions': safe_float(c(59)),
        'audience_ctr': safe_float(c(60)),
        'audience_visitors': safe_float(c(61)),
        'audience_visitor_ratio': safe_float(c(62)),
        'audience_ppc': safe_float(c(63)),
        'audience_conversion_cost': safe_float(c(64)),
        'audience_cart_qty': safe_float(c(65)),
        'audience_cart_rate': safe_float(c(66)),
        'audience_cart_cost': safe_float(c(67)),
        'full_site_ad_spend': safe_float(c(68)),
        'full_site_ad_sales': safe_float(c(69)),
        'full_site_ad_roi': safe_float(c(70)),
        'full_site_ad_conversion': safe_float(c(71)),
        'full_site_impressions': safe_float(c(72)),
        'full_site_ctr': safe_float(c(73)),
        'full_site_visitors': safe_float(c(74)),
        'full_site_visitor_ratio': safe_float(c(75)),
        'full_site_ppc': safe_float(c(76)),
        'full_site_conversion_cost': safe_float(c(77)),
        'full_site_cart_qty': safe_float(c(78)),
        'full_site_cart_rate': safe_float(c(79)),
        'full_site_cart_cost': safe_float(c(80)),
    }


def map_row_monthly_summary(ws, row, col_start=1):
    """Map row for product_monthly_summary table only."""
    def c(offset):
        return ws.cell(row=row, column=col_start + offset - 1).value
    return {
        'source_file': '',
        'month': '',
        'product_id': extract_pid(c(1)),
        'product_title': c(2),
        'image_url': c(3),
        'category': c(4),
        'list_date': c(5),
        'payment_amount': safe_float(c(6)),
        'refund_amount': safe_float(c(7)),
        'net_sales': safe_float(c(8)),
        'ad_spend': safe_float(c(9)),
        'ad_roi': safe_float(c(10)),
        'refund_paid_ratio': safe_float(c(11)),
        'paid_ratio': safe_float(c(12)),
        'visitors': safe_int(c(13)),
        'uv_value': safe_float(c(14)),
        'payment_conversion': safe_float(c(15)),
        'refund_rate': safe_float(c(16)),
        'cart_rate': safe_float(c(17)),
        'fav_rate': safe_float(c(18)),
        'bounce_rate': safe_float(c(19)),
        'avg_stay_duration': safe_float(c(20)),
        'search_conversion': safe_float(c(21)),
        'score': safe_float(c(22)),
        'page_views': safe_int(c(23)),
        'payment_buyers': safe_int(c(24)),
        'aov': safe_float(c(25)),
        'payment_qty': safe_int(c(26)),
        'cart_qty': safe_int(c(27)),
        'fav_users': safe_int(c(28)),
        'search_buyers': safe_int(c(29)),
        'search_ratio': safe_float(c(30)),
        'search_uv_value': safe_float(c(31)),
        'total_orders': safe_int(c(32)),
        'store_favs': safe_int(c(33)),
        'total_favs': safe_int(c(34)),
        'fav_cost': safe_float(c(35)),
        'fav_cart_qty': safe_int(c(36)),
        'order_conversion': safe_float(c(37)),
        'ctr': safe_float(c(38)),
        'total_ad_spend': safe_float(c(39)),
        'impressions': safe_int(c(40)),
        'clicks': safe_int(c(41)),
        'cart_items': safe_int(c(42)),
        'keyword_ad_spend': safe_float(c(43)),
        'keyword_ad_sales': safe_float(c(44)),
        'keyword_ad_roi': safe_float(c(45)),
        'keyword_ad_conversion': safe_float(c(46)),
        'keyword_impressions': safe_int(c(47)),
        'keyword_ctr': safe_float(c(48)),
        'keyword_visitors': safe_int(c(49)),
        'keyword_visitor_ratio': safe_float(c(50)),
        'keyword_ppc': safe_float(c(51)),
        'keyword_conversion_cost': safe_float(c(52)),
        'keyword_cart_qty': safe_int(c(53)),
        'keyword_cart_rate': safe_float(c(54)),
        'keyword_cart_cost': safe_float(c(55)),
        'audience_ad_spend': safe_float(c(56)),
        'audience_ad_sales': safe_float(c(57)),
        'audience_ad_roi': safe_float(c(58)),
        'audience_ad_conversion': safe_float(c(59)),
        'audience_impressions': safe_int(c(60)),
        'audience_ctr': safe_float(c(61)),
        'audience_visitors': safe_int(c(62)),
        'audience_visitor_ratio': safe_float(c(63)),
        'audience_ppc': safe_float(c(64)),
        'audience_conversion_cost': safe_float(c(65)),
        'audience_cart_qty': safe_int(c(66)),
        'audience_cart_rate': safe_float(c(67)),
        'audience_cart_cost': safe_float(c(68)),
        'full_site_ad_spend': safe_float(c(69)),
        'full_site_ad_sales': safe_float(c(70)),
        'full_site_ad_roi': safe_float(c(71)),
        'full_site_ad_conversion': safe_float(c(72)),
        'full_site_impressions': safe_int(c(73)),
        'full_site_ctr': safe_float(c(74)),
        'full_site_visitors': safe_int(c(75)),
        'full_site_visitor_ratio': safe_float(c(76)),
        'full_site_ppc': safe_float(c(77)),
        'full_site_conversion_cost': safe_float(c(78)),
        'full_site_cart_qty': safe_int(c(79)),
        'full_site_cart_rate': safe_float(c(80)),
        'full_site_cart_cost': safe_float(c(81)),
    }

def insert_planning(cursor, plan_month, d):
    if not d['product_id'] or d['product_id'] == '0':
        return False
    cursor.execute("DELETE FROM monthly_planning WHERE plan_month=? AND product_id=?", (plan_month, d['product_id']))
    cols = list(d.keys())
    vals = list(d.values())
    placeholders = ','.join(['?' for _ in cols])
    col_names = ','.join(cols)
    sql = f"INSERT INTO monthly_planning (plan_month, {col_names}) VALUES (?, {placeholders})"
    cursor.execute(sql, [plan_month] + vals)
    return True

def import_monthly_planning(db):
    cursor = db.cursor()
    if not os.path.exists(PLANNING_FILE):
        print(f"[WARN] 规划文件不存在: {PLANNING_FILE}")
        return
    
    print(f"\n[导入] 5月规划: {PLANNING_FILE}")
    wb = openpyxl.load_workbook(PLANNING_FILE, data_only=True)
    total = 0
    
    if '5月单品规划' in wb.sheetnames:
        ws = wb['5月单品规划']
        print(f"  [处理] 5月单品规划 (行1={ws.max_row})...")
        for row in range(4, ws.max_row + 1):
            pid = extract_pid(ws.cell(row=row, column=5).value)
            if not pid or pid == '0':
                continue
            d = {
                'plan_month': '5月',
                'product_id': pid,
                'product_title': ws.cell(row=row, column=6).value,
                'image_url': ws.cell(row=row, column=8).value,
                'category': ws.cell(row=row, column=1).value,
                'tier': ws.cell(row=row, column=4).value,
                'manager': ws.cell(row=row, column=2).value,
                'new_old': ws.cell(row=row, column=3).value,
                'list_date': ws.cell(row=row, column=7).value,
            }
            if insert_planning(cursor, '5月', d):
                total += 1
        print(f"  [OK] 插入 {total} 条5月规划基础数据")
    
    if '4月单品' in wb.sheetnames:
        ws = wb['4月单品']
        print(f"  [处理] 4月单品...")
        month = '4月'
        cnt = 0
        for row in range(2, ws.max_row + 1):
            d = map_row_planning(ws, row, col_start=1)
            if insert_planning(cursor, month, d):
                cnt += 1
        total += cnt
        print(f"  [OK] 插入 {cnt} 条4月单品数据")
    
    if '5月单品' in wb.sheetnames:
        ws = wb['5月单品']
        print(f"  [处理] 5月单品 (日数据)...")
        month = '5月'
        cnt = 0
        for row in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value or ''
            d = map_row_planning(ws, row, col_start=2)
            d['plan_month'] = str(date_val)[:10] if date_val else month
            if insert_planning(cursor, d['plan_month'], d):
                cnt += 1
        total += cnt
        print(f"  [OK] 插入 {cnt} 条5月单品日数据")
    
    wb.close()
    db.commit()
    print(f"  [总计] 插入 {total} 条月度规划数据")

def import_monthly_data(db):
    cursor = db.cursor()
    if not os.path.exists(MONTHLY_FILE):
        print(f"[WARN] 月度数据文件不存在: {MONTHLY_FILE}")
        return
    
    print(f"\n[导入] 月度数据分析: {MONTHLY_FILE}")
    wb = openpyxl.load_workbook(MONTHLY_FILE, data_only=True)
    
    # DMP-源
    if 'DMP-源' in wb.sheetnames:
        ws = wb['DMP-源']
        print(f"  [处理] DMP-源...")
        total = 0
        for row in range(3, ws.max_row + 1):
            pid = extract_pid(ws.cell(row=row, column=1).value)
            if not pid:
                continue
            cursor.execute("DELETE FROM dmp_product_data WHERE product_id=?", (pid,))
            cursor.execute("""INSERT INTO dmp_product_data (
                product_id, product_title, growth_stage, payment_amount, ipv, ad_ipv, ad_cost, ad_roi,
                cart_fav_rate, payment_conversion, repurchase_rate, presale_amount, presale_qty,
                organic_ipv, search_ipv, recommend_ipv, search_ctr, unit_price,
                cross_sell_qty, cross_sell_rate, cross_sell_categories, repurchase_users
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                pid, ws.cell(row=row, column=2).value, ws.cell(row=row, column=3).value,
                safe_float(ws.cell(row=row, column=4).value), safe_int(ws.cell(row=row, column=5).value),
                safe_int(ws.cell(row=row, column=6).value), safe_float(ws.cell(row=row, column=7).value),
                safe_float(ws.cell(row=row, column=8).value), safe_float(ws.cell(row=row, column=9).value),
                safe_float(ws.cell(row=row, column=10).value), safe_float(ws.cell(row=row, column=11).value),
                safe_float(ws.cell(row=row, column=12).value), safe_int(ws.cell(row=row, column=13).value),
                safe_int(ws.cell(row=row, column=14).value), safe_int(ws.cell(row=row, column=15).value),
                safe_int(ws.cell(row=row, column=16).value), safe_float(ws.cell(row=row, column=17).value),
                safe_float(ws.cell(row=row, column=18).value), safe_int(ws.cell(row=row, column=19).value),
                safe_float(ws.cell(row=row, column=20).value), safe_int(ws.cell(row=row, column=21).value),
                safe_int(ws.cell(row=row, column=22).value),
            ))
            total += 1
        print(f"  [OK] DMP数据: {total} 条")
    
    # 付费-源
    if '付费-源' in wb.sheetnames:
        ws = wb['付费-源']
        print(f"  [处理] 付费-源...")
        total = 0
        for row in range(2, ws.max_row + 1):
            eid = extract_pid(ws.cell(row=row, column=2).value)
            if not eid:
                continue
            vals = [ws.cell(row=row, column=c).value for c in range(1, 76)]
            cursor.execute("""INSERT INTO paid_source_data (
                date, entity_id, entity_type, entity_name,
                impressions, clicks, cost, ctr, cpc, cpm,
                presale_gmv, presale_orders, direct_presale_gmv, direct_presale_orders,
                indirect_presale_gmv, indirect_presale_orders, direct_gmv, indirect_gmv,
                total_gmv, total_orders, direct_orders, indirect_orders,
                click_conversion, roi, roi_with_presale, order_cost,
                total_cart, direct_cart, indirect_cart, cart_rate,
                fav_items, fav_store, store_fav_cost, total_fav_cart, total_fav_cart_cost,
                item_fav_cart, item_fav_cart_cost, total_favs, item_fav_cost, item_fav_rate, cart_cost,
                order_qty, order_amount,
                wangwang_consult, guide_visits, guide_visitors, guide_potential, guide_potential_ratio,
                member_rate, member_count, guide_visit_rate, deep_visits, avg_pages,
                new_buyers, new_buyer_ratio, member_first_buy, member_gmv, member_orders,
                buyers, orders_per_buyer, amount_per_buyer,
                organic_conversion, organic_exposure, platform_boost_gmv, platform_boost_direct,
                platform_boost_clicks
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                normalize_date(vals[0]), eid, vals[2], vals[3],
                safe_int(vals[4]), safe_int(vals[5]), safe_float(vals[6]), safe_float(vals[7]), safe_float(vals[8]), safe_float(vals[9]),
                safe_float(vals[10]), safe_int(vals[11]), safe_float(vals[12]), safe_int(vals[13]),
                safe_float(vals[14]), safe_int(vals[15]), safe_float(vals[16]), safe_float(vals[17]),
                safe_float(vals[18]), safe_int(vals[19]), safe_int(vals[20]), safe_int(vals[21]),
                safe_float(vals[22]), safe_float(vals[23]), safe_float(vals[24]), safe_float(vals[25]),
                safe_int(vals[26]), safe_int(vals[27]), safe_int(vals[28]), safe_float(vals[29]),
                safe_int(vals[30]), safe_int(vals[31]), safe_float(vals[32]), safe_int(vals[33]), safe_float(vals[34]),
                safe_int(vals[35]), safe_float(vals[36]), safe_int(vals[37]), safe_float(vals[38]), safe_float(vals[39]), safe_float(vals[40]),
                safe_int(vals[41]), safe_float(vals[42]),
                safe_int(vals[48]), safe_int(vals[49]), safe_int(vals[50]), safe_int(vals[51]), safe_float(vals[52]),
                safe_float(vals[53]), safe_int(vals[54]), safe_float(vals[55]), safe_int(vals[56]), safe_float(vals[57]),
                safe_int(vals[58]), safe_float(vals[59]), safe_int(vals[60]), safe_float(vals[61]), safe_int(vals[62]),
                safe_int(vals[63]), safe_float(vals[64]), safe_float(vals[65]),
                safe_float(vals[66]), safe_int(vals[67]), safe_float(vals[68]), safe_float(vals[69]),
                safe_int(vals[70]),
            ))
            total += 1
        print(f"  [OK] 付费数据: {total} 条")
    
    # 单品总表-源
    if '单品总表-源' in wb.sheetnames:
        ws = wb['单品总表-源']
        print(f"  [处理] 单品总表-源...")
        total = 0
        for row in range(2, ws.max_row + 1):
            pid = extract_pid(ws.cell(row=row, column=3).value)
            if not pid:
                continue
            month_val = str(ws.cell(row=row, column=2).value or '')
            d = map_row_monthly_summary(ws, row, col_start=7)
            d['source_file'] = str(ws.cell(row=row, column=1).value or '')[:50]
            d['month'] = month_val
            cols = list(d.keys())
            vals = list(d.values())
            ph = ','.join(['?' for _ in cols])
            cn = ','.join(cols)
            cursor.execute(f"INSERT INTO product_monthly_summary ({cn}) VALUES ({ph})", vals)
            total += 1
        print(f"  [OK] 单品汇总: {total} 条")
    
    # 销售-源
    if '销售-源' in wb.sheetnames:
        ws = wb['销售-源']
        print(f"  [处理] 销售-源...")
        total = 0
        cols_list = [
            'date', 'category',
            'payment_amount', 'store_customers', 'avg_stay_duration', 'payment_buyers', 'payment_conversion',
            'aov', 'repurchase_rate', 'repurchase_buyers', 'repurchase_amount',
            'payment_orders', 'payment_qty', 'visitors', 'page_views',
            'refund_amount', 'refund_rate', 'net_payment', 'cart_qty', 'cart_users', 'fav_users',
            'consult_rate', 'wangwang_response_time', 'shipping_rate_24h', 'delivery_time_hours',
            'refund_process_days', 'refund_success_rate', 'platform_dispute_rate',
            'full_site_ad_spend', 'keyword_ad_spend', 'audience_ad_spend', 'scene_ad_spend',
            'taoke_commission', 'total_cost',
            'old_buyers', 'revisit_buyers', 'consult_users', 'content_viewers', 'past_buyers',
            'wangwang_satisfaction', 'total_payment', 'shipping_rate_48h', 'related_products',
            'content_scale', 'dispute_rate', 'product_coverage', 'broadcast_hours',
            'consult_conversion', 'old_buyer_amount', 'public_video_count', 'total_orders', 'stay_duration',
        ]
        placeholders = ','.join(['?' for _ in cols_list])
        col_names = ','.join(cols_list)
        sql = f"INSERT INTO sales_source_monthly ({col_names}) VALUES ({placeholders})"
        for row in range(2, ws.max_row + 1):
            date_val = normalize_date(ws.cell(row=row, column=1).value)
            category = str(ws.cell(row=row, column=2).value or '')
            vals = [safe_float(ws.cell(row=row, column=c).value) for c in range(3, 53)]
            cursor.execute(sql, [date_val, category] + vals)
            total += 1
        print(f"  [OK] 销售源数据: {total} 条")
    
    wb.close()
    db.commit()

def main():
    print("="*60)
    print("生命周期数据导入脚本")
    print("="*60)
    print(f"数据库: {DB_PATH}")
    
    db = sqlite3.connect(DB_PATH)
    try:
        import_monthly_planning(db)
        import_monthly_data(db)
        
        print("\n" + "="*60)
        print("导入统计:")
        cursor = db.cursor()
        for t in ['monthly_planning', 'dmp_product_data', 'paid_source_data', 'product_monthly_summary', 'sales_source_monthly']:
            try:
                count = cursor.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                print(f"  {t}: {count} rows")
            except:
                print(f"  {t}: ERROR")
    finally:
        db.close()
    
    print("\n" + "="*60)
    print("[SUCCESS] 导入完成!")

if __name__ == '__main__':
    main()
