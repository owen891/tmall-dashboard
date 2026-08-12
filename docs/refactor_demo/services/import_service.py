"""
Import Service — 数据导入编排。

从 scripts/import_data.py 和 data_api.py 的 /api/upload/data 路由提取。
负责 Sheet 识别 → 数据清洗 → 入库的编排流程。
"""
import os
from datetime import datetime
from openpyxl import load_workbook

from models import db
from models.product import Product
from models.data import DailyData, WeeklyData, MonthlyData
from models.paid import PaidDetail


class ImportService:

    # Sheet 名称 → 数据类型映射
    SHEET_MAP = {
        '商品数据': 'monthly',
        '月度数据': 'monthly',
        '周度数据': 'weekly',
        '日度数据': 'daily',
        '推广数据': 'paid',
        '付费明细': 'paid',
    }

    @staticmethod
    def import_workbook(file_path, progress_callback=None):
        """
        导入 Excel 工作簿 — 替代原 import_data.py 的主流程。

        流程：
        1. 打开工作簿，识别 Sheet
        2. 按 Sheet 类型分发到对应的清洗+入库方法
        3. 返回导入结果摘要
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(file_path, data_only=True, read_only=True)
        results = {'sheets': [], 'total_rows': 0, 'warnings': []}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data_type = ImportService._identify_sheet(sheet_name, ws)

            if data_type is None:
                results['warnings'].append(f"Unknown sheet: {sheet_name}, skipped")
                continue

            if progress_callback:
                progress_callback(sheet_name, data_type)

            count = ImportService._import_sheet(ws, data_type)
            results['sheets'].append({
                'sheet': sheet_name,
                'type': data_type,
                'rows': count,
            })
            results['total_rows'] += count

        wb.close()
        return results

    @staticmethod
    def _identify_sheet(sheet_name, ws):
        """识别 Sheet 类型"""
        # 先按名称匹配
        for name, dtype in ImportService.SHEET_MAP.items():
            if name in sheet_name:
                return dtype

        # 按表头列名推断
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_set = {str(h).strip() for h in headers if h}

        if '月份' in header_set or 'month' in header_set:
            return 'monthly'
        if '周开始' in header_set or 'week_start' in header_set:
            return 'weekly'
        if '日期' in header_set and 'product_id' in header_set:
            return 'daily'
        if '花费' in header_set or 'impressions' in header_set:
            return 'paid'

        return None

    @staticmethod
    def _import_sheet(ws, data_type):
        """按类型导入 Sheet"""
        importers = {
            'monthly': ImportService._import_monthly,
            'weekly': ImportService._import_weekly,
            'daily': ImportService._import_daily,
            'paid': ImportService._import_paid,
        }
        importer = importers.get(data_type)
        if not importer:
            return 0
        return importer(ws)

    @staticmethod
    def _import_monthly(ws):
        """导入月度数据"""
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        count = 0
        for row in rows:
            if not row or not row[0]:
                continue
            product_id = str(row[0]).strip()

            # 确保商品存在
            ImportService._ensure_product(product_id, row)

            # upsert 月度数据
            record = MonthlyData.query.filter_by(
                product_id=product_id, month=str(row[1]).strip()
            ).first()

            if record:
                ImportService._update_record(record, row)
            else:
                ImportService._create_monthly(product_id, row)

            count += 1

        db.session.commit()
        return count

    @staticmethod
    def _import_weekly(ws):
        """导入周度数据 — 结构类似 _import_monthly"""
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        count = 0
        for row in rows:
            if not row or not row[0]:
                continue
            product_id = str(row[0]).strip()
            ImportService._ensure_product(product_id, row)
            # ... 同月度逻辑
            count += 1
        db.session.commit()
        return count

    @staticmethod
    def _import_daily(ws):
        """导入日度数据"""
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        count = 0
        for row in rows:
            if not row or not row[0]:
                continue
            product_id = str(row[0]).strip()
            ImportService._ensure_product(product_id, row)
            # ... 同月度逻辑
            count += 1
        db.session.commit()
        return count

    @staticmethod
    def _import_paid(ws):
        """导入付费推广数据"""
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        count = 0
        for row in rows:
            if not row or not row[0]:
                continue
            product_id = str(row[0]).strip()
            date_range = str(row[1]).strip()

            record = PaidDetail.query.filter_by(
                product_id=product_id, date_range=date_range
            ).first()

            if not record:
                record = PaidDetail(product_id=product_id, date_range=date_range)
                db.session.add(record)

            count += 1

        db.session.commit()
        return count

    @staticmethod
    def _ensure_product(product_id, row):
        """确保商品记录存在"""
        existing = Product.query.filter_by(product_id=product_id).first()
        if not existing:
            # 从行数据中提取商品信息
            title = str(row[2]).strip() if len(row) > 2 and row[2] else product_id
            product = Product(product_id=product_id, title=title)
            db.session.add(product)
            db.session.flush()  # 获取 ID

    @staticmethod
    def _create_monthly(product_id, row):
        """创建月度数据记录"""
        record = MonthlyData(
            product_id=product_id,
            month=str(row[1]).strip(),
        )
        ImportService._update_record(record, row)
        db.session.add(record)

    @staticmethod
    def _update_record(record, row):
        """根据行数据更新记录字段（通用方法）"""
        # 此处根据实际 Excel 列顺序映射字段
        # 示例：假设 row[2]=payment_amount, row[3]=refund_amount, ...
        field_map = {
            2: 'payment_amount', 3: 'refund_amount', 4: 'net_sales',
            5: 'visitors', 6: 'payment_conversion', 7: 'cart_rate',
            8: 'ad_spend', 9: 'ad_roi', 10: 'refund_rate',
        }
        for idx, field in field_map.items():
            if idx < len(row) and row[idx] is not None:
                try:
                    setattr(record, field, float(row[idx]))
                except (ValueError, TypeError):
                    pass
